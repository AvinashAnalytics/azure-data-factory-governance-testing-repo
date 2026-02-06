"""Excel Enhancement Module - Professional Excel formatting and styling capabilities"""

from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    numbers, Color, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, CellIsRule, Rule
)
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.comments import Comment
from typing import Any, Dict, List, Tuple, Optional
import re
import traceback
from pathlib import Path
from datetime import datetime
from collections import Counter
import json

class EnhancementConfig:
    """
     MODULAR ENHANCEMENT CONFIGURATION

    Allows enabling/disabling features individually
    """

    DEFAULT_CONFIG = {
        "excel_enhancements": {
            "enabled": True,
            "core_formatting": {
                "enabled": True,
                "column_sizing": True,
                "number_format": True,
                "alignment": True,
                "borders": True,
                "row_shading": True,
                "header_style": True
            },
            "conditional_formatting": {
                "enabled": True,
                "data_bars": True,
                "icon_sets": True,
                "color_scales": True,
                "status_highlighting": True
            },
            "hyperlinks": {
                "enabled": True,
                "summary_navigation": True,
                "auto_convert_references": True
            },
            "protection": {
                "enabled": False,
                "password": None
            },
            "enhanced_summary": {
                "enabled": True,
                "project_banner": True,
                "executive_summary": True,
                "critical_alerts": True,
                "metrics_dashboard": True,
                "resource_overview": True,
                "recommendations": True
            },
            "advanced_dashboard": {
                "enabled": True,
                "health_score": True,
                "cost_analysis": False,
                "complexity_heat_map": True,
                "performance_insights": True,
                "top_pipelines": True,
                "security_checklist": True,
                "activity_distribution": True,
                "network_stats": True,
                "change_risk": True
            },
            "page_setup": {
                "enabled": True,
                "orientation": "landscape"
            }
        }
    }

    @staticmethod
    def load_config(config_file: str = "enhancement_config.json") -> Dict:
        """
        Load enhancement configuration from file

        Args:
            config_file: Path to config file

        Returns:
            Configuration dictionary
        """
        # Primary: relative to current working directory
        config_path = Path(config_file)
        # Fallback: next to this module (works with dashboard UI save location)
        module_fallback = Path(__file__).parent / config_file

        for path in (config_path, module_fallback):
            if path.exists():
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        config = json.load(f)
                    print(f" Loaded enhancement config from: {path}")
                    return config
                except Exception as e:
                    print(f"  Config file error at {path}, using defaults: {e}")
                    return EnhancementConfig.DEFAULT_CONFIG

        print("‚Ñπ  No config file found, using default settings")
        return EnhancementConfig.DEFAULT_CONFIG

    @staticmethod
    def is_enabled(config: Dict, *path) -> bool:
        """
        Check if a feature is enabled

        Args:
            config: Configuration dictionary
            *path: Path to feature (e.g., 'core_formatting', 'column_sizing')

        Returns:
            True if enabled, False otherwise
        """
        current = config.get('excel_enhancements', {})

        for key in path:
            if isinstance(current, dict):
                current = current.get(key, False)
            else:
                return False

        return bool(current)

ENHANCEMENT_CONFIG = EnhancementConfig.load_config()

class ExcelTheme:
    """
     Modern Professional Excel Theme

    Color Palette inspired by Microsoft Fluent Design
    """

    HEADER_BG = "2F5496"           # Professional Blue
    HEADER_TEXT = "FFFFFF"         # White

    ROW_EVEN = "FFFFFF"            # White
    ROW_ODD = "F2F2F2"             # Light Gray

    CRITICAL = "C00000"            # Dark Red
    HIGH = "FF6600"                # Orange
    MEDIUM = "FFC000"              # Amber
    LOW = "92D050"                 # Light Green
    SUCCESS = "00B050"             # Green
    WARNING = "FFFF00"             # Yellow
    INFO = "00B0F0"                # Light Blue

    ORPHANED_BG = "FFF2CC"         # Light Yellow
    ERROR_BG = "FFE6E6"            # Light Red
    SUMMARY_BG = "E7F3FF"          # Very Light Blue

    BORDER_DARK = "404040"         # Dark Gray
    BORDER_LIGHT = "D0D0D0"        # Light Gray

    TEXT_PRIMARY = "000000"        # Black
    TEXT_SECONDARY = "595959"      # Medium Gray

    TEXT_LINK = "0563C1"
    TEXT_LINK = "0563C1"           # Blue (hyperlink)

class ExcelBorders:
    """ Pre-defined border styles"""

    @staticmethod
    def thin_border(color=ExcelTheme.BORDER_LIGHT):
        """Thin border for data cells"""
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def thick_border(color=ExcelTheme.BORDER_DARK):
        """Thick border for headers"""
        side = Side(style='medium', color=color)
        return Border(left=side, right=side, top=side, bottom=side)

    @staticmethod
    def header_border():
        """Special border for header row"""
        thin = Side(style='thin', color=ExcelTheme.BORDER_LIGHT)
        thick = Side(style='medium', color=ExcelTheme.BORDER_DARK)
        return Border(left=thin, right=thin, top=thin, bottom=thick)

class IntelligentColumnSizer:
    """
     INTELLIGENT COLUMN WIDTH CALCULATOR

    Features:
    - Content-aware sizing
    - Special handling for URLs, SQL, JSON
    - Header consideration
    - Multi-line text detection
    - Optimal width algorithms
    """

    MIN_WIDTH = 8
    MAX_WIDTH = 100
    DEFAULT_WIDTH = 12

    WIDTH_SHORT_CODE = 10      # Status, Type codes
    WIDTH_NAME = 30            # Resource names
    WIDTH_DESCRIPTION = 50     # Descriptions
    WIDTH_SQL = 80             # SQL queries
    WIDTH_URL = 60             # URLs, file paths
    WIDTH_COUNT = 12           # Numeric counts
    WIDTH_PERCENTAGE = 10      # Percentages
    WIDTH_DATE = 20            # Dates/timestamps

    @classmethod
    def calculate_column_width(cls, column_cells, header_name: str = "") -> int:
        """
         Calculate optimal column width

        Args:
            column_cells: List of cells in column
            header_name: Column header name for type detection

        Returns:
            Optimal width (between MIN_WIDTH and MAX_WIDTH)
        """

        col_type = cls._detect_column_type(header_name)

        if col_type:
            base_width = cls._get_type_width(col_type)

            if col_type in ['count', 'percentage', 'status']:
                return base_width
        else:
            base_width = cls.DEFAULT_WIDTH

        max_length = 0
        has_multiline = False

        for cell in column_cells:
            if cell.value is None:
                continue

            cell_value = str(cell.value)

            if '\n' in cell_value:
                has_multiline = True
                lines = cell_value.split('\n')
                cell_length = max(len(line) for line in lines)
            else:
                cell_length = len(cell_value)

            if cls._is_url(cell_value):
                cell_length = min(cell_length, cls.WIDTH_URL)
            elif cls._is_sql(cell_value):
                cell_length = min(cell_length, cls.WIDTH_SQL)
            elif cls._is_json(cell_value):
                cell_length = min(cell_length, cls.WIDTH_DESCRIPTION)

            max_length = max(max_length, cell_length)

        calculated_width = max_length + 2

        final_width = max(cls.MIN_WIDTH, min(calculated_width, cls.MAX_WIDTH))

        if base_width and abs(final_width - base_width) < 5:
            final_width = base_width

        if has_multiline:
            final_width = min(final_width, cls.WIDTH_DESCRIPTION)

        return final_width

    @classmethod
    def _detect_column_type(cls, header: str) -> str:
        """Detect column type from header name"""
        if not header:
            return ""

        header_lower = header.lower()

        if any(x in header_lower for x in ['status', 'state', 'type', 'level', 'severity', 'impact']):
            return 'status'

        if any(x in header_lower for x in ['count', 'total', 'number', 'depth', 'sequence']):
            return 'count'

        if 'percentage' in header_lower or header_lower.endswith('%'):
            return 'percentage'

        if any(x in header_lower for x in ['date', 'time', 'timestamp', 'created', 'modified']):
            return 'date'

        if any(x in header_lower for x in ['name', 'pipeline', 'dataset', 'activity', 'trigger']):
            return 'name'

        if any(x in header_lower for x in ['description', 'details', 'message', 'reason']):
            return 'description'

        if any(x in header_lower for x in ['sql', 'query', 'script', 'command']):
            return 'sql'

        if any(x in header_lower for x in ['url', 'path', 'file', 'location', 'link']):
            return 'url'

        return ""

    @classmethod
    def _get_type_width(cls, col_type: str) -> int:
        """Get recommended width for column type"""
        type_widths = {
            'status': cls.WIDTH_SHORT_CODE,
            'count': cls.WIDTH_COUNT,
            'percentage': cls.WIDTH_PERCENTAGE,
            'date': cls.WIDTH_DATE,
            'name': cls.WIDTH_NAME,
            'description': cls.WIDTH_DESCRIPTION,
            'sql': cls.WIDTH_SQL,
            'url': cls.WIDTH_URL
        }
        return type_widths.get(col_type, cls.DEFAULT_WIDTH)

    @staticmethod
    def _is_url(text: str) -> bool:
        """Check if text is a URL"""
        return text.startswith(('http://', 'https://', 'ftp://', '//'))

    @staticmethod
    def _is_sql(text: str) -> bool:
        """Check if text is SQL"""
        sql_keywords = ['SELECT', 'INSERT', 'UPDATE', 'DELETE', 'CREATE', 'ALTER', 'DROP', 'EXEC']
        text_upper = text.upper()
        return any(text_upper.startswith(kw) for kw in sql_keywords)

    @staticmethod
    def _is_json(text: str) -> bool:
        """Check if text is JSON"""
        return (text.startswith('{') and text.endswith('}')) or \
               (text.startswith('[') and text.endswith(']'))

class NumberFormatter:
    """
     INTELLIGENT NUMBER FORMATTING

    Applies appropriate number formats based on column content
    """

    FORMAT_INTEGER = '#,##0'                    # 1,234
    FORMAT_DECIMAL = '#,##0.00'                 # 1,234.56
    FORMAT_PERCENTAGE = '0.0%'                  # 45.5%
    FORMAT_PERCENTAGE_INT = '0%'                # 45%
    FORMAT_CURRENCY = '$#,##0.00'               # $1,234.56
    FORMAT_DATE = 'yyyy-mm-dd'                  # 2024-01-15
    FORMAT_DATETIME = 'yyyy-mm-dd hh:mm:ss'     # 2024-01-15 14:30:00
    FORMAT_TIME = 'hh:mm:ss'                    # 14:30:00

    @classmethod
    def apply_number_format(cls, worksheet, header_row: int = 1):
        """
         Apply number formatting to entire worksheet

        Args:
            worksheet: openpyxl worksheet
            header_row: Row number of headers (1-based)
        """

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for col_idx, header in headers.items():
            col_letter = get_column_letter(col_idx)

            number_format = cls._detect_format(header)

            if number_format:

                for row in range(header_row + 1, worksheet.max_row + 1):
                    cell = worksheet[f'{col_letter}{row}']

                    if cell.value is not None:

                        if 'percentage' in header or header.endswith('%'):
                            if isinstance(cell.value, (int, float)):

                                if cell.value > 1:
                                    cell.value = cell.value / 100

                        cell.number_format = number_format

    @classmethod
    def _detect_format(cls, header: str) -> str:
        """Detect appropriate number format from header"""

        if 'percentage' in header or header.endswith('%'):
            return cls.FORMAT_PERCENTAGE

        if any(x in header for x in ['count', 'total', 'number', 'depth', 'sequence']):
            return cls.FORMAT_INTEGER

        if any(x in header for x in ['cost', 'price', 'amount', 'fee']):
            return cls.FORMAT_CURRENCY

        if 'date' in header and 'update' not in header:
            return cls.FORMAT_DATE

        if any(x in header for x in ['timestamp', 'datetime', 'created', 'modified']):
            return cls.FORMAT_DATETIME

        if 'time' in header and 'runtime' not in header:
            return cls.FORMAT_TIME

        if any(x in header for x in ['score', 'rating', 'average', 'mean']):
            return cls.FORMAT_DECIMAL

        return ""

class CellAlignmentManager:
    """
     INTELLIGENT CELL ALIGNMENT

    Applies professional alignment based on content type
    """

    @staticmethod
    def apply_alignment(worksheet, header_row: int = 1):
        """
         Apply intelligent alignment to worksheet

        Rules:
        - Headers: Center + Bold
        - Numbers: Right align
        - Text: Left align
        - Long text: Left + Wrap
        """

        for cell in worksheet[header_row]:
            if cell.value:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            for col_idx, header in headers.items():
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f'{col_letter}{row_idx}']

                if cell.value is None:
                    continue

                h_align, wrap = CellAlignmentManager._get_alignment(header, cell.value)

                cell.alignment = Alignment(
                    horizontal=h_align,
                    vertical='top',
                    wrap_text=wrap
                )

    @staticmethod
    def _get_alignment(header: str, value: Any) -> Tuple[str, bool]:
        """
        Determine horizontal alignment and wrap setting

        Returns:
            Tuple of (horizontal_alignment, wrap_text)
        """

        if isinstance(value, (int, float)):
            return ('right', False)

        if any(x in header for x in ['count', 'total', 'number', 'percentage', '%']):
            return ('right', False)

        value_str = str(value)

        if len(value_str) > 50:
            return ('left', True)

        if any(x in header for x in ['sql', 'query', 'description', 'details', 'reason', 'message']):
            return ('left', True)

        if any(x in header for x in ['status', 'state', 'type', 'impact', 'severity', 'level']):
            return ('center', False)

        return ('left', False)

class BorderApplier:
    """
     PROFESSIONAL BORDER APPLICATION

    Adds clean, professional borders to all cells
    """

    @staticmethod
    def apply_borders(worksheet, header_row: int = 1):
        """
         Apply professional borders

        - Header row: Thick bottom border
        - Data cells: Thin borders
        - Alternating row shading for readability
        """

        for cell in worksheet[header_row]:
            cell.border = ExcelBorders.header_border()

        for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = ExcelBorders.thin_border()

class AlternatingRowShader:
    """
     ALTERNATING ROW SHADING

    Makes large tables easier to read
    """

    @staticmethod
    def apply_shading(worksheet, header_row: int = 1):
        """
         Apply alternating row colors

        Even rows: White
        Odd rows: Light gray
        """

        even_fill = PatternFill(
            start_color=ExcelTheme.ROW_EVEN,
            end_color=ExcelTheme.ROW_EVEN,
            fill_type='solid'
        )

        odd_fill = PatternFill(
            start_color=ExcelTheme.ROW_ODD,
            end_color=ExcelTheme.ROW_ODD,
            fill_type='solid'
        )

        for row_idx in range(header_row + 1, worksheet.max_row + 1):

            fill = even_fill if (row_idx - header_row) % 2 == 0 else odd_fill

            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row_idx, col_idx)

                if not cell.fill or cell.fill.start_color.rgb == '00000000':
                    cell.fill = fill

class MasterFormatter:
    """
     MASTER FORMATTER

    Orchestrates all formatting operations in optimal order
    """

    @staticmethod
    def format_worksheet(worksheet, sheet_name: str = "", header_row: int = 1,
                        enable_features: Dict[str, bool] = None):
        """
         Apply complete professional formatting to worksheet

        Args:
            worksheet: openpyxl worksheet
            sheet_name: Name of sheet (for special handling)
            header_row: Row number of headers
            enable_features: Dict to enable/disable features
                {
                    'column_sizing': True,
                    'number_format': True,
                    'alignment': True,
                    'borders': True,
                    'row_shading': True,
                    'header_style': True
                }
        """

        if enable_features is None:
            enable_features = {
                'column_sizing': True,
                'number_format': True,
                'alignment': True,
                'borders': True,
                'row_shading': True,
                'header_style': True
            }

        try:

            if enable_features.get('column_sizing', True):
                MasterFormatter._apply_column_sizing(worksheet, header_row)

            if enable_features.get('number_format', True):
                NumberFormatter.apply_number_format(worksheet, header_row)

            if enable_features.get('alignment', True):
                CellAlignmentManager.apply_alignment(worksheet, header_row)

            if enable_features.get('row_shading', True):
                AlternatingRowShader.apply_shading(worksheet, header_row)

            if enable_features.get('borders', True):
                BorderApplier.apply_borders(worksheet, header_row)

            if enable_features.get('header_style', True):
                MasterFormatter._apply_header_style(worksheet, header_row)

        except Exception as e:
            print(f"  Warning: Formatting failed for {sheet_name}: {e}")

    @staticmethod
    def _apply_column_sizing(worksheet, header_row: int):
        """Apply intelligent column sizing"""
        for column in worksheet.columns:
            col_letter = get_column_letter(column[0].column)

            header_cell = worksheet[f'{col_letter}{header_row}']
            header_name = str(header_cell.value) if header_cell.value else ""

            width = IntelligentColumnSizer.calculate_column_width(column, header_name)

            worksheet.column_dimensions[col_letter].width = width

    @staticmethod
    def _apply_header_style(worksheet, header_row: int):
        """Apply professional header styling"""

        header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=ExcelTheme.HEADER_TEXT
        )

        header_fill = PatternFill(
            start_color=ExcelTheme.HEADER_BG,
            end_color=ExcelTheme.HEADER_BG,
            fill_type='solid'
        )

        for cell in worksheet[header_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill

print(" Part 1/6 loaded: Core Enhancement Framework")
"""Excel Enhancement Module - Professional Excel formatting and styling capabilities"""

class DataBarFormatter:
    """
     DATA BAR FORMATTER

    Adds visual progress bars to numeric columns
    Perfect for: counts, usage statistics, percentages
    """

    BLUE_GRADIENT = {
        'color': "4472C4",      # Professional Blue
        'border_color': "2F5496"
    }

    GREEN_GRADIENT = {
        'color': "70AD47",      # Success Green
        'border_color': "548235"
    }

    ORANGE_GRADIENT = {
        'color': "ED7D31",      # Warning Orange
        'border_color': "C65911"
    }

    RED_GRADIENT = {
        'color': "E74C3C",      # Alert Red
        'border_color': "C0392B"
    }

    @staticmethod
    def add_data_bars(worksheet, column_letter: str, start_row: int, end_row: int,
                     color_scheme: Dict = None, show_value: bool = True):
        """
         Add data bars to a column

        Args:
            worksheet: openpyxl worksheet
            column_letter: Column letter (e.g., 'C')
            start_row: First data row
            end_row: Last data row
            color_scheme: Color scheme dict (default: blue)
            show_value: Show numeric value alongside bar
        """

        if color_scheme is None:
            color_scheme = DataBarFormatter.BLUE_GRADIENT

        cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

        data_bar = DataBarRule(
            start_type='min',
            start_value=None,
            end_type='max',
            end_value=None,
            color=color_scheme['color'],
            showValue=show_value,
            minLength=0,
            maxLength=100
        )

        worksheet.conditional_formatting.add(cell_range, data_bar)

    @staticmethod
    def auto_add_data_bars(worksheet, header_row: int = 1):
        """
         Automatically add data bars to appropriate columns

        Detects columns that should have data bars:
        - Count columns
        - Usage columns
        - Numeric metrics
        """

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for col_idx, header in headers.items():

            if DataBarFormatter._should_have_data_bar(header):

                col_letter = get_column_letter(col_idx)
                start_row = header_row + 1
                end_row = worksheet.max_row

                if end_row < start_row:
                    continue

                color_scheme = DataBarFormatter._get_color_scheme(header)

                DataBarFormatter.add_data_bars(
                    worksheet, col_letter, start_row, end_row,
                    color_scheme=color_scheme,
                    show_value=True
                )

    @staticmethod
    def _should_have_data_bar(header: str) -> bool:
        """Check if column should have data bars"""

        data_bar_keywords = [
            'count', 'total', 'usage', 'number', 'activities',
            'references', 'consumers', 'depth', 'blastradius'
        ]

        if 'percentage' in header or header.endswith('%'):
            return False

        return any(keyword in header for keyword in data_bar_keywords)

    @staticmethod
    def _get_color_scheme(header: str) -> Dict:
        """Get appropriate color scheme for column"""

        if any(x in header for x in ['error', 'warning', 'orphaned', 'broken']):
            return DataBarFormatter.RED_GRADIENT

        if any(x in header for x in ['usage', 'used', 'success', 'complete']):
            return DataBarFormatter.GREEN_GRADIENT

        if any(x in header for x in ['pending', 'medium', 'depth']):
            return DataBarFormatter.ORANGE_GRADIENT

        return DataBarFormatter.BLUE_GRADIENT

class IconSetFormatter:
    """
     ICON SET FORMATTER

    Adds visual indicators (traffic lights, arrows, flags)
    Perfect for: status, impact levels, severity
    """

    TRAFFIC_LIGHTS = "3TrafficLights1"      # 
    ARROWS = "3Arrows"                       # ‚Üì‚Üí‚Üë
    FLAGS = "3Flags"                         # üö©üè≥üèÅ
    SYMBOLS = "3Symbols"                     # ‚úó‚óã‚úì
    STARS = "3Stars"                         # ‚òÜ‚òÖ‚òÖ
    TRIANGLES = "3Triangles"                 # ‚ñΩ‚ñ≥‚ñ≤

    @staticmethod
    def add_icon_set(worksheet, column_letter: str, start_row: int, end_row: int,
                    icon_style: str = None, reverse: bool = False):
        """
         Add icon set to a column

        Args:
            worksheet: openpyxl worksheet
            column_letter: Column letter
            start_row: First data row
            end_row: Last data row
            icon_style: Icon set style (default: traffic lights)
            reverse: Reverse icon order (green=low, red=high)
        """

        if icon_style is None:
            icon_style = IconSetFormatter.TRAFFIC_LIGHTS

        cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

        icon_set = IconSetRule(
            icon_style=icon_style,
            type='percent',
            values=[33, 67],
            showValue=True,
            reverse=reverse
        )

        worksheet.conditional_formatting.add(cell_range, icon_set)

    @staticmethod
    def auto_add_icon_sets(worksheet, header_row: int = 1):
        """
         Automatically add icon sets to appropriate columns

        Detects columns that should have icons:
        - Status columns
        - Impact/Severity columns
        - Complexity columns
        """

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for col_idx, header in headers.items():

            icon_config = IconSetFormatter._get_icon_config(header)

            if icon_config:
                col_letter = get_column_letter(col_idx)
                start_row = header_row + 1
                end_row = worksheet.max_row

                if end_row < start_row:
                    continue

                IconSetFormatter.add_icon_set(
                    worksheet, col_letter, start_row, end_row,
                    icon_style=icon_config['style'],
                    reverse=icon_config['reverse']
                )

    @staticmethod
    def _get_icon_config(header: str) -> Optional[Dict]:
        """Get icon configuration for column"""

        if 'complexity' in header:
            return {
                'style': IconSetFormatter.TRIANGLES,
                'reverse': True  # Green for low complexity
            }

        if any(x in header for x in ['impact', 'severity', 'priority']):
            return {
                'style': IconSetFormatter.TRAFFIC_LIGHTS,
                'reverse': False  # Red for high impact
            }

        if 'status' in header or 'state' in header:
            return {
                'style': IconSetFormatter.SYMBOLS,
                'reverse': False
            }

        if 'depth' in header or 'level' in header:
            return {
                'style': IconSetFormatter.ARROWS,
                'reverse': False  # Up arrow for high depth
            }

        return None

class ColorScaleFormatter:
    """
     COLOR SCALE FORMATTER (HEAT MAPS)

    Adds gradient color scales to show data distribution
    Perfect for: percentages, scores, metrics
    """

    RED_YELLOW_GREEN = {
        'min_color': "F8696B",   # Red
        'mid_color': "FFEB84",   # Yellow
        'max_color': "63BE7B"    # Green
    }

    WHITE_RED = {
        'min_color': "FFFFFF",   # White
        'max_color': "F8696B"    # Red
    }

    WHITE_BLUE = {
        'min_color': "FFFFFF",   # White
        'max_color': "5A8AC6"    # Blue
    }

    GREEN_YELLOW_RED = {
        'min_color': "63BE7B",   # Green
        'mid_color': "FFEB84",   # Yellow
        'max_color': "F8696B"    # Red
    }

    @staticmethod
    def add_color_scale(worksheet, column_letter: str, start_row: int, end_row: int,
                       color_scale: Dict = None, use_midpoint: bool = True):
        """
         Add color scale to a column

        Args:
            worksheet: openpyxl worksheet
            column_letter: Column letter
            start_row: First data row
            end_row: Last data row
            color_scale: Color scale dict
            use_midpoint: Use 3-color scale (True) or 2-color (False)
        """

        if color_scale is None:
            color_scale = ColorScaleFormatter.RED_YELLOW_GREEN

        cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

        if use_midpoint and 'mid_color' in color_scale:

            rule = ColorScaleRule(
                start_type='min',
                start_color=color_scale['min_color'],
                mid_type='percentile',
                mid_value=50,
                mid_color=color_scale['mid_color'],
                end_type='max',
                end_color=color_scale['max_color']
            )
        else:

            rule = ColorScaleRule(
                start_type='min',
                start_color=color_scale['min_color'],
                end_type='max',
                end_color=color_scale['max_color']
            )

        worksheet.conditional_formatting.add(cell_range, rule)

    @staticmethod
    def auto_add_color_scales(worksheet, header_row: int = 1):
        """
         Automatically add color scales to appropriate columns

        Detects columns that should have color scales:
        - Percentage columns
        - Score columns
        - Complexity columns
        """

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for col_idx, header in headers.items():

            scale_config = ColorScaleFormatter._get_scale_config(header)

            if scale_config:
                col_letter = get_column_letter(col_idx)
                start_row = header_row + 1
                end_row = worksheet.max_row

                if end_row < start_row:
                    continue

                ColorScaleFormatter.add_color_scale(
                    worksheet, col_letter, start_row, end_row,
                    color_scale=scale_config['colors'],
                    use_midpoint=scale_config['use_midpoint']
                )

    @staticmethod
    def _get_scale_config(header: str) -> Optional[Dict]:
        """Get color scale configuration for column"""

        if 'percentage' in header or header.endswith('%'):
            return {
                'colors': ColorScaleFormatter.WHITE_BLUE,
                'use_midpoint': False
            }

        if 'complexity' in header and 'score' in header:
            return {
                'colors': ColorScaleFormatter.GREEN_YELLOW_RED,
                'use_midpoint': True
            }

        if any(x in header for x in ['performance', 'efficiency', 'quality']):
            return {
                'colors': ColorScaleFormatter.RED_YELLOW_GREEN,
                'use_midpoint': True
            }

        return None

class StatusFormatter:
    """
     STATUS-BASED CONDITIONAL FORMATTING

    Highlights cells based on specific text values
    Perfect for: Impact levels, Severity, Status columns
    """

    @staticmethod
    def add_status_highlighting(worksheet, column_letter: str, start_row: int, end_row: int,
                               status_colors: Dict[str, str]):
        """
         Add status-based highlighting

        Args:
            worksheet: openpyxl worksheet
            column_letter: Column letter
            start_row: First data row
            end_row: Last data row
            status_colors: Dict mapping status values to colors
                Example: {'CRITICAL': 'FF0000', 'HIGH': 'FFA500', ...}
        """

        cell_range = f"{column_letter}{start_row}:{column_letter}{end_row}"

        for status_value, color in status_colors.items():

            fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type='solid'
            )

            if StatusFormatter._is_dark_color(color):
                font = Font(color="FFFFFF", bold=True)
            else:
                font = Font(color="000000", bold=True)

            rule = CellIsRule(
                operator='equal',
                formula=[f'"{status_value}"'],
                fill=fill,
                font=font
            )

            worksheet.conditional_formatting.add(cell_range, rule)

    @staticmethod
    def auto_add_status_highlighting(worksheet, header_row: int = 1):
        """
         Automatically add status highlighting to appropriate columns

        Detects and applies highlighting to:
        - Impact columns (CRITICAL, HIGH, MEDIUM, LOW)
        - Severity columns
        - Status columns (Started, Stopped, Success, Failed)
        - Yes/No columns
        """

        headers = {}
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                headers[col_idx] = str(cell.value).lower()

        for col_idx, header in headers.items():

            status_colors = StatusFormatter._get_status_colors(header)

            if status_colors:
                col_letter = get_column_letter(col_idx)
                start_row = header_row + 1
                end_row = worksheet.max_row

                if end_row < start_row:
                    continue

                StatusFormatter.add_status_highlighting(
                    worksheet, col_letter, start_row, end_row,
                    status_colors=status_colors
                )

    @staticmethod
    def _get_status_colors(header: str) -> Optional[Dict[str, str]]:
        """Get status-to-color mapping for column"""

        if 'impact' in header:
            return {
                'CRITICAL': ExcelTheme.CRITICAL,
                'HIGH': ExcelTheme.HIGH,
                'MEDIUM': ExcelTheme.MEDIUM,
                'LOW': ExcelTheme.LOW
            }

        if 'severity' in header:
            return {
                'CRITICAL': ExcelTheme.CRITICAL,
                'HIGH': ExcelTheme.HIGH,
                'MEDIUM': ExcelTheme.MEDIUM,
                'LOW': ExcelTheme.LOW
            }

        if 'complexity' in header:
            return {
                'Critical': ExcelTheme.CRITICAL,
                'High': ExcelTheme.HIGH,
                'Medium': ExcelTheme.MEDIUM,
                'Low': ExcelTheme.LOW
            }

        if 'status' in header or 'state' in header:
            return {
                'Started': ExcelTheme.SUCCESS,
                'Stopped': ExcelTheme.WARNING,
                'Running': ExcelTheme.SUCCESS,
                'Failed': ExcelTheme.CRITICAL,
                'Success': ExcelTheme.SUCCESS,
                'Error': ExcelTheme.CRITICAL
            }

        if 'orphaned' in header or 'broken' in header or 'isorphaned' in header:
            return {
                'Yes': ExcelTheme.WARNING,
                'No': ExcelTheme.SUCCESS
            }

        if 'multi' in header or 'ismulti' in header:
            return {
                'Yes': ExcelTheme.INFO,
                'No': ExcelTheme.ROW_ODD
            }

        return None

    @staticmethod
    def _is_dark_color(hex_color: str) -> bool:
        """Check if color is dark (needs white text)"""

        hex_color = hex_color.lstrip('#')

        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        luminance = (0.299 * r + 0.587 * g + 0.114 * b)

        return luminance < 128

class MasterConditionalFormatter:
    """
     MASTER CONDITIONAL FORMATTER

    Orchestrates all conditional formatting operations
    """

    @staticmethod
    def apply_all_conditional_formatting(worksheet, sheet_name: str = "",
                                        header_row: int = 1,
                                        enable_features: Dict[str, bool] = None):
        """
         Apply all conditional formatting to worksheet

        Args:
            worksheet: openpyxl worksheet
            sheet_name: Name of sheet (for special handling)
            header_row: Row number of headers
            enable_features: Dict to enable/disable features
        """

        if enable_features is None:
            enable_features = {
                'data_bars': True,
                'icon_sets': True,
                'color_scales': True,
                'status_highlighting': True
            }

        try:

            if worksheet.max_row <= header_row:
                return

            if enable_features.get('status_highlighting', True):
                StatusFormatter.auto_add_status_highlighting(worksheet, header_row)

            if enable_features.get('data_bars', True):
                DataBarFormatter.auto_add_data_bars(worksheet, header_row)

            if enable_features.get('icon_sets', True):

                IconSetFormatter.auto_add_icon_sets(worksheet, header_row)

            if enable_features.get('color_scales', True):
                ColorScaleFormatter.auto_add_color_scales(worksheet, header_row)

        except Exception as e:
            print(f"  Warning: Conditional formatting failed for {sheet_name}: {e}")

class SpecialSheetFormatters:
    """
     SPECIAL FORMATTERS FOR SPECIFIC SHEETS

    Custom formatting for Summary, ImpactAnalysis, etc.
    """

    @staticmethod
    def format_summary_sheet(worksheet, header_row: int = 1):
        """
         Special formatting for Summary sheet

        - Highlights critical issues
        - Color-codes metrics
        - Emphasizes important values
        """

        value_col = None
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value and 'value' in str(cell.value).lower():
                value_col = col_idx
                break

        if value_col:
            col_letter = get_column_letter(value_col)

            DataBarFormatter.add_data_bars(
                worksheet, col_letter, header_row + 1, worksheet.max_row,
                color_scheme=DataBarFormatter.BLUE_GRADIENT,
                show_value=True
            )

    @staticmethod
    def format_impact_analysis_sheet(worksheet, header_row: int = 1):
        """
         Special formatting for ImpactAnalysis sheet

        - Impact level highlighting
        - Blast radius visualization
        - Dependency count indicators
        """

        pass

    @staticmethod
    def format_circular_dependencies_sheet(worksheet, header_row: int = 1):
        """
         Special formatting for CircularDependencies sheet

        - Highlights entire rows for CRITICAL severity
        - Color-codes cycle length
        """

        severity_col = None
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value and 'severity' in str(cell.value).lower():
                severity_col = col_idx
                break

        if severity_col:

            for row_idx in range(header_row + 1, worksheet.max_row + 1):
                severity_cell = worksheet.cell(row_idx, severity_col)

                if severity_cell.value == 'CRITICAL':

                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row_idx, col_idx)
                        cell.fill = PatternFill(
                            start_color=ExcelTheme.ERROR_BG,
                            end_color=ExcelTheme.ERROR_BG,
                            fill_type='solid'
                        )
                        cell.font = Font(bold=True)

print(" Part 2/6 loaded: Enhanced Conditional Formatting")
"""Excel Enhancement Module - Professional Excel formatting and styling capabilities"""

class HyperlinkManager:
    """
     HYPERLINK MANAGER

     CRITICAL FIX: Converts text references to clickable hyperlinks

    Creates:
    - Internal sheet navigation links
    - External URL links
    - Email links
    - Professional link styling
    """

    LINK_FONT = Font(
        name='Calibri',
        size=11,
        underline='single',
        color=ExcelTheme.TEXT_LINK
    )

    @staticmethod
    def create_internal_link(worksheet, cell, target_sheet: str, display_text: str = None):
        """
         Create clickable internal sheet link

        Args:
            worksheet: Source worksheet
            cell: Cell to add hyperlink to
            target_sheet: Target sheet name
            display_text: Text to display (default: current cell value)
        """

        if display_text is None:
            display_text = str(cell.value) if cell.value else target_sheet

        escaped_sheet = target_sheet.replace("'", "''")

        if ' ' in target_sheet or any(c in target_sheet for c in ['!', '#', '$']):
            link_formula = f"#'{escaped_sheet}'!A1"
        else:
            link_formula = f"#{escaped_sheet}!A1"

        cell.hyperlink = link_formula

        cell.value = display_text

        cell.font = HyperlinkManager.LINK_FONT

    @staticmethod
    def create_external_link(worksheet, cell, url: str, display_text: str = None):
        """
         Create clickable external URL link

        Args:
            worksheet: Worksheet
            cell: Cell to add hyperlink to
            url: Full URL (must start with http://, https://, etc.)
            display_text: Text to display (default: URL)
        """

        if display_text is None:
            display_text = url

        cell.hyperlink = url

        cell.value = display_text

        cell.font = HyperlinkManager.LINK_FONT

    @staticmethod
    def create_email_link(worksheet, cell, email: str, subject: str = "",
                         display_text: str = None):
        """
         Create clickable email link

        Args:
            worksheet: Worksheet
            cell: Cell to add hyperlink to
            email: Email address
            subject: Email subject (optional)
            display_text: Text to display (default: email)
        """

        if display_text is None:
            display_text = email

        if subject:
            link = f"mailto:{email}?subject={subject}"
        else:
            link = f"mailto:{email}"

        cell.hyperlink = link

        cell.value = display_text

        cell.font = HyperlinkManager.LINK_FONT

    @staticmethod
    def auto_convert_sheet_references(worksheet, available_sheets: List[str],
                                     header_row: int = 1):
        """
         CRITICAL FIX: Auto-convert text sheet references to hyperlinks

        Scans "Details" column for patterns like:
        - " See sheet: PipelineAnalysis"
        - "See sheet: Activities"
        - "Sheet: DataFlows"

        And converts them to clickable links.

        Args:
            worksheet: Worksheet to process
            available_sheets: List of all sheet names in workbook
            header_row: Row number of headers
        """

        details_col = None
        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value and 'details' in str(cell.value).lower():
                details_col = col_idx
                break

        if not details_col:
            return

        col_letter = get_column_letter(details_col)

        sheet_lookup = {sheet.lower(): sheet for sheet in available_sheets}

        patterns = [
            r'\s*See(?:\s*sheet)?:\s*([^,;\n]+)',
            r'See(?:\s*sheet)?:\s*([^,;\n]+)',
            r'Sheet:\s*([^,;\n]+)',
            r'‚Üí\s*([^,;\n]+)',
        ]

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet[f'{col_letter}{row_idx}']

            if not cell.value:
                continue

            cell_text = str(cell.value)

            for pattern in patterns:
                match = re.search(pattern, cell_text, re.IGNORECASE)

                if match:

                    mentioned_sheet = match.group(1).strip().strip('"\'').strip()

                    mentioned_sheet = re.sub(r'\s+', ' ', mentioned_sheet)

                    actual_sheet = sheet_lookup.get(mentioned_sheet.lower())

                    if actual_sheet:

                        display_text = str(cell.value).strip()
                        HyperlinkManager.create_internal_link(
                            worksheet, cell, actual_sheet, display_text
                        )
                        break

    @staticmethod
    def add_navigation_links_to_summary(summary_worksheet, all_sheets: List[str]):
        """
         CRITICAL FIX: Add navigation section to Summary sheet

        Creates a "Quick Navigation" section at the bottom with links to all sheets.

        Args:
            summary_worksheet: Summary worksheet
            all_sheets: List of all sheet names
        """

        last_row = summary_worksheet.max_row

        nav_start_row = last_row + 3

        header_cell = summary_worksheet.cell(nav_start_row, 1)
        header_cell.value = "QUICK NAVIGATION"
        header_cell.font = Font(name='Calibri', size=14, bold=True, color=ExcelTheme.HEADER_TEXT)
        header_cell.fill = PatternFill(
            start_color=ExcelTheme.HEADER_BG,
            end_color=ExcelTheme.HEADER_BG,
            fill_type='solid'
        )

        desc_cell = summary_worksheet.cell(nav_start_row, 2)
        desc_cell.value = "Click links below to navigate to sheets"
        desc_cell.font = Font(name='Calibri', size=10, italic=True)

        current_row = nav_start_row + 2

        categorized_sheets = HyperlinkManager._categorize_sheets(all_sheets)

        for category, sheets in categorized_sheets.items():
            if not sheets:
                continue

            cat_cell = summary_worksheet.cell(current_row, 1)
            cat_cell.value = f"üìÅ {category}"
            cat_cell.font = Font(bold=True, size=11)
            current_row += 1

            for sheet_name in sheets:

                name_cell = summary_worksheet.cell(current_row, 1)
                name_cell.value = f"  ‚Ä¢ {sheet_name}"

                link_cell = summary_worksheet.cell(current_row, 2)
                HyperlinkManager.create_internal_link(
                    summary_worksheet, link_cell, sheet_name, f"‚Üí Open {sheet_name}"
                )

                current_row += 1

            current_row += 1

    @staticmethod
    def _categorize_sheets(sheets: List[str]) -> Dict[str, List[str]]:
        """Categorize sheets for better navigation"""

        categories = {
            'Overview': [],
            'Pipelines': [],
            'Activities': [],
            'DataFlows': [],
            'Resources': [],
            'Analysis': [],
            'Orphaned Resources': [],
            'Statistics': [],
            'Other': []
        }

        for sheet in sheets:
            sheet_lower = sheet.lower()

            if sheet_lower == 'summary':
                continue

            if any(x in sheet_lower for x in ['pipeline']):
                categories['Pipelines'].append(sheet)
            elif any(x in sheet_lower for x in ['activity', 'activities']):
                categories['Activities'].append(sheet)
            elif any(x in sheet_lower for x in ['dataflow', 'lineage', 'transformation']):
                categories['DataFlows'].append(sheet)
            elif any(x in sheet_lower for x in ['dataset', 'linkedservice', 'trigger', 'integration']):
                categories['Resources'].append(sheet)
            elif any(x in sheet_lower for x in ['impact', 'circular', 'dependency']):
                categories['Analysis'].append(sheet)
            elif 'orphaned' in sheet_lower:
                categories['Orphaned Resources'].append(sheet)
            elif any(x in sheet_lower for x in ['usage', 'count', 'statistics']):
                categories['Statistics'].append(sheet)
            else:
                categories['Other'].append(sheet)

        return {k: v for k, v in categories.items() if v}

class ExcelTableFormatter:
    """
     EXCEL TABLE FORMATTER

    Converts ranges to Excel Tables with professional styling

    Benefits:
    - Auto-filter built-in
    - Professional appearance
    - Easy sorting
    - Named ranges
    """

    STYLE_BLUE = "TableStyleMedium2"
    STYLE_ORANGE = "TableStyleMedium4"
    STYLE_GREEN = "TableStyleMedium3"
    STYLE_LIGHT = "TableStyleLight1"

    @staticmethod
    def create_table(worksheet, table_name: str, ref: str, style: str = None,
                    show_row_stripes: bool = True):
        """
         Create Excel Table from range

        Args:
            worksheet: Worksheet
            table_name: Unique table name
            ref: Cell range (e.g., "A1:F100")
            style: Table style (default: blue)
            show_row_stripes: Show alternating row colors
        """

        if style is None:
            style = ExcelTableFormatter.STYLE_BLUE

        table = Table(displayName=table_name, ref=ref)

        table_style = TableStyleInfo(
            name=style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=show_row_stripes,
            showColumnStripes=False
        )
        table.tableStyleInfo = table_style

        worksheet.add_table(table)

    @staticmethod
    def auto_create_table(worksheet, sheet_name: str, header_row: int = 1):
        """
         Auto-create Excel Table for entire data range

        Args:
            worksheet: Worksheet
            sheet_name: Sheet name (used for table naming)
            header_row: Header row number
        """

        if worksheet.max_row <= header_row:
            return

        last_col = get_column_letter(worksheet.max_column)
        ref = f"A{header_row}:{last_col}{worksheet.max_row}"

        table_name = re.sub(r'[^a-zA-Z0-9_]', '_', sheet_name)
        table_name = f"Table_{table_name}"

        style = ExcelTableFormatter._get_table_style(sheet_name)

        try:

            ExcelTableFormatter.create_table(
                worksheet, table_name, ref, style=style
            )
        except Exception as e:

            print(f"  Could not create table for {sheet_name}: {e}")

    @staticmethod
    def _get_table_style(sheet_name: str) -> str:
        """Get appropriate table style for sheet"""

        sheet_lower = sheet_name.lower()

        if any(x in sheet_lower for x in ['error', 'circular', 'orphaned']):
            return ExcelTableFormatter.STYLE_ORANGE

        if any(x in sheet_lower for x in ['usage', 'success']):
            return ExcelTableFormatter.STYLE_GREEN

        return ExcelTableFormatter.STYLE_LIGHT

class SheetProtectionManager:
    """
     SHEET PROTECTION MANAGER

    Protects sheets while allowing filtering and selection

    Benefits:
    - Prevents accidental edits
    - Allows filtering
    - Allows sorting
    - Professional workbook
    """

    @staticmethod
    def protect_sheet(worksheet, password: str = None,
                     allow_filter: bool = True,
                     allow_sort: bool = True,
                     allow_select_locked: bool = True,
                     allow_select_unlocked: bool = True):
        """
         Protect sheet with user-friendly settings

        Args:
            worksheet: Worksheet to protect
            password: Optional password (default: no password)
            allow_filter: Allow auto-filter
            allow_sort: Allow sorting
            allow_select_locked: Allow selecting locked cells
            allow_select_unlocked: Allow selecting unlocked cells
        """

        protection = SheetProtection(
            sheet=True,
            password=password,
            autoFilter=allow_filter,
            sort=allow_sort,
            selectLockedCells=allow_select_locked,
            selectUnlockedCells=allow_select_unlocked,
            formatCells=False,
            formatColumns=False,
            formatRows=False,
            insertColumns=False,
            insertRows=False,
            deleteColumns=False,
            deleteRows=False
        )

        worksheet.protection = protection

    @staticmethod
    def protect_all_sheets(workbook, sheets_to_protect: List[str] = None,
                          password: str = None):
        """
         Protect multiple sheets in workbook

        Args:
            workbook: openpyxl Workbook
            sheets_to_protect: List of sheet names (default: all except Summary)
            password: Optional password
        """

        if sheets_to_protect is None:

            sheets_to_protect = [
                ws.title for ws in workbook.worksheets
                if ws.title.lower() != 'summary'
            ]

        for sheet_name in sheets_to_protect:
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                SheetProtectionManager.protect_sheet(
                    worksheet, password=password
                )

class CellCommentManager:
    """
     CELL COMMENT MANAGER

    Adds helpful tooltips and documentation to cells
    """

    @staticmethod
    def add_comment(worksheet, cell, text: str, author: str = "ADF Analyzer"):
        """
         Add comment/tooltip to cell

        Args:
            worksheet: Worksheet
            cell: Cell to add comment to
            text: Comment text
            author: Comment author
        """

        comment = Comment(text, author)
        cell.comment = comment

    @staticmethod
    def add_header_comments(worksheet, header_descriptions: Dict[str, str],
                           header_row: int = 1):
        """
         Add helpful comments to column headers

        Args:
            worksheet: Worksheet
            header_descriptions: Dict mapping header names to descriptions
            header_row: Header row number
        """

        for col_idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value:
                header_name = str(cell.value)

                if header_name in header_descriptions:
                    CellCommentManager.add_comment(
                        worksheet, cell, header_descriptions[header_name]
                    )

    @staticmethod
    def auto_add_helpful_comments(worksheet, sheet_name: str, header_row: int = 1):
        """
         Automatically add helpful comments to common columns

        Args:
            worksheet: Worksheet
            sheet_name: Sheet name (for context-aware comments)
            header_row: Header row number
        """

        descriptions = CellCommentManager._get_standard_descriptions()

        sheet_descriptions = CellCommentManager._get_sheet_specific_descriptions(sheet_name)
        descriptions.update(sheet_descriptions)

        CellCommentManager.add_header_comments(worksheet, descriptions, header_row)

    @staticmethod
    def _get_standard_descriptions() -> Dict[str, str]:
        """Get standard column descriptions"""

        return {
            'Impact': 'Impact level based on dependencies:\nCRITICAL = High upstream+downstream\nHIGH = Significant dependencies\nMEDIUM = Entry point\nLOW = Orphaned/standalone',
            'BlastRadius': 'Total number of resources affected by changes to this resource',
            'Complexity': 'Complexity assessment:\nCritical = 100+\nHigh = 50-99\nMedium = 20-49\nLow = <20',
            'IsOrphaned': 'Yes = Not referenced by any trigger or active pipeline\nNo = Actively used',
            'Sequence': 'Execution order within pipeline (lower numbers execute first)',
            'Depth': 'Nesting level (0=root, higher=more nested)',
            'IntegrationRuntime': 'Runtime used for execution:\nAutoResolveIR = Azure auto-managed\nOther = Self-hosted or custom IR',
            'UsageCount': 'Number of times this resource is referenced',
            'State': 'Trigger state:\nStarted = Active\nStopped = Inactive',
            'Type': 'Resource type classification'
        }

    @staticmethod
    def _get_sheet_specific_descriptions(sheet_name: str) -> Dict[str, str]:
        """Get sheet-specific column descriptions"""

        sheet_lower = sheet_name.lower()

        if 'impact' in sheet_lower:
            return {
                'DirectUpstreamTriggers': 'Triggers that directly invoke this pipeline',
                'TransitiveUpstreamPipelines': 'Pipelines in the dependency chain (up to 5 levels)',
                'DirectDownstreamPipelines': 'Pipelines directly called by this pipeline',
                'TransitiveDownstreamPipelines': 'All downstream pipelines in chain'
            }

        elif 'circular' in sheet_lower:
            return {
                'Cycle': 'Dependency cycle path (A‚ÜíB‚ÜíC‚ÜíA)',
                'Length': 'Number of resources in cycle',
                'Severity': 'CRITICAL = Production blocker, must fix immediately'
            }

        elif 'orphaned' in sheet_lower:
            return {
                'Reason': 'Why this resource is considered orphaned',
                'Recommendation': 'Suggested action to resolve orphan status'
            }

        return {}

class PageSetupManager:
    """
     PAGE SETUP MANAGER

    Configures print settings for professional output
    """

    @staticmethod
    def setup_page(worksheet, orientation: str = 'landscape',
                  paper_size: int = 9,  # 9 = A4
                  fit_to_width: int = 1,
                  fit_to_height: int = 0,  # 0 = unlimited
                  header_text: str = None,
                  footer_text: str = None):
        """
         Setup page for printing

        Args:
            worksheet: Worksheet
            orientation: 'landscape' or 'portrait'
            paper_size: Paper size code (9=A4, 1=Letter)
            fit_to_width: Number of pages wide (1=fit to 1 page)
            fit_to_height: Number of pages tall (0=unlimited)
            header_text: Custom header text
            footer_text: Custom footer text
        """

        worksheet.page_setup.orientation = orientation
        worksheet.page_setup.paperSize = paper_size
        worksheet.page_setup.fitToWidth = fit_to_width
        worksheet.page_setup.fitToHeight = fit_to_height

        worksheet.page_margins = PageMargins(
            left=0.7, right=0.7,
            top=0.75, bottom=0.75,
            header=0.3, footer=0.3
        )

        if header_text:
            worksheet.oddHeader.center.text = header_text
        else:
            worksheet.oddHeader.center.text = f"&A"  # Sheet name

        if footer_text:
            worksheet.oddFooter.center.text = footer_text
        else:
            worksheet.oddFooter.left.text = "ADF Analyzer v10.0"
            worksheet.oddFooter.center.text = "Page &P of &N"
            worksheet.oddFooter.right.text = "&D"  # Date

        worksheet.print_options.horizontalCentered = True
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def auto_setup_all_sheets(workbook):
        """
         Auto-setup all sheets for printing

        Args:
            workbook: openpyxl Workbook
        """

        for worksheet in workbook.worksheets:
            try:
                PageSetupManager.setup_page(
                    worksheet,
                    orientation='landscape',
                    header_text=f"{worksheet.title}",
                    footer_text=None
                )
            except Exception as e:
                print(f"  Page setup failed for {worksheet.title}: {e}")

print(" Part 3/6 loaded: Hyperlinks, Protection & Advanced Features")

"""Excel Enhancement Module - Professional Excel formatting and styling capabilities"""

def create_enhanced_export_function(analyzer_class):
    """
     CREATE ENHANCED EXPORT FUNCTION

    This replaces the original export_to_excel() with beautified version
    """

    original_export = analyzer_class.export_to_excel

    def enhanced_export_to_excel(self):
        """
         ENHANCED EXCEL EXPORT WITH ALL BEAUTIFICATION

        Includes:
        - All original functionality
        - Intelligent column sizing
        - Professional formatting
        - Conditional formatting
        - Hyperlinks in Summary
        - Sheet protection
        - Print settings
        """

        # Respect runtime config: if excel enhancements are disabled, call original export
        try:
            master_enabled = ENHANCEMENT_CONFIG.get('excel_enhancements', {}).get('enabled', True)
        except Exception:
            master_enabled = True

        if not master_enabled:
            try:
                self.logger.info("Excel enhancements disabled via config; delegating to original export")
            except Exception:
                pass
            return original_export(self)

        from datetime import datetime
        import shutil
        from pathlib import Path
        import pandas as pd

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = Path('output')
        output_dir.mkdir(exist_ok=True)

        excel_file = output_dir / 'adf_analysis_latest.xlsx'
        archive_file = output_dir / f'adf_analysis_{timestamp}.xlsx'

        self.logger.info(f" Exporting to Excel with ENHANCED BEAUTIFICATION: {excel_file}")

        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:

                self._used_sheet_names = set()

                self._write_summary_sheet(writer, timestamp)

                # Normalize activity result records: ensure new columns exist before writing sheets
                try:
                    for a in self.results.get('activities', []):
                        # Preserve backwards-compatibility while adding clearer names
                        a['ParseSequence'] = a.get('Sequence', '')
                        a['HasDependsOn'] = 'Yes' if a.get('Dependencies') else 'No'
                        a['DependsOnCount'] = len(a.get('Dependencies') or [])
                        a['CycleFlag'] = 'Yes' if a.get('ExecutionStage') == 'CYCLE' else 'No'
                except Exception:
                    pass

                # Dump small debug sample of results before writing core sheets
                try:
                    import json, os
                    os.makedirs('output', exist_ok=True)
                    sample = {
                        'activities_sample': self.results.get('activities', [])[:5],
                        'activity_execution_order_sample': self.results.get('activity_execution_order', [])[:5]
                    }
                    with open('output/debug_stage_samples.json', 'w', encoding='utf-8') as fh:
                        json.dump(sample, fh, default=str, indent=2)
                except Exception:
                    pass

                self._write_core_data_sheets(writer)

                # Ensure ExecutionStage headers exist in the in-memory workbook
                try:
                    wb = writer.book
                    if 'Activities' in wb.sheetnames:
                        ws = wb['Activities']
                        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                        # Ensure activity-level new columns exist
                        for col in ('ParseSequence','ExecutionStage','HasDependsOn','DependsOnCount','CycleFlag'):
                            if col not in headers:
                                ws.cell(row=1, column=len(headers) + 1, value=col)
                                headers.append(col)
                    if 'ActivityExecutionOrder' in wb.sheetnames:
                        ws2 = wb['ActivityExecutionOrder']
                        headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
                        if 'FromExecutionStage' not in headers2:
                            ws2.cell(row=1, column=len(headers2) + 1, value='FromExecutionStage')
                            headers2.append('FromExecutionStage')
                        if 'ToExecutionStage' not in headers2:
                            ws2.cell(row=1, column=len(headers2) + 1, value='ToExecutionStage')
                            headers2.append('ToExecutionStage')
                except Exception:
                    # non-critical; continue with normal flow
                    pass

                self._write_analysis_sheets(writer)

                self._write_orphaned_sheets(writer)

                self._write_usage_statistics_sheets(writer)

                self._write_additional_resource_sheets(writer)

                self._write_errors_sheet(writer)

                self.logger.info(" Applying enhanced beautification...")
                self._apply_enhanced_beautification(writer)

            self.logger.info(f" Export complete with BEAUTIFICATION: {excel_file}")

            # Post-process: rewrite Activities and ActivityExecutionOrder from in-memory results
            try:
                import pandas as pd
                from openpyxl import load_workbook
                from pathlib import Path

                act_df = pd.DataFrame(self.results.get('activities', []))
                order_df = pd.DataFrame(self.results.get('activity_execution_order', []))

                # Load hide_config to remove hidden columns before writing
                hidden_columns = {}
                try:
                    for cfg_path in ['enhancement_config.json', 'config/enhancement_config.json']:
                        if Path(cfg_path).exists():
                            with open(cfg_path, 'r') as f:
                                hide_cfg = json.load(f).get('hide_config', {})
                                if hide_cfg.get('enabled', False):
                                    hidden_columns = hide_cfg.get('hidden_columns', {})
                                break
                except Exception:
                    pass
                
                # Remove hidden columns from DataFrames
                activities_hidden = hidden_columns.get('Activities', [])
                if activities_hidden and not act_df.empty:
                    act_df = act_df.drop(columns=[c for c in activities_hidden if c in act_df.columns], errors='ignore')
                
                aeo_hidden = hidden_columns.get('ActivityExecutionOrder', [])
                if aeo_hidden and not order_df.empty:
                    order_df = order_df.drop(columns=[c for c in aeo_hidden if c in order_df.columns], errors='ignore')

                # Sort DataFrames for clean Excel output
                # Activities: sort by Pipeline name A-Z, then ExecutionStage smallest to largest
                if not act_df.empty and 'Pipeline' in act_df.columns:
                    sort_cols_act = ['Pipeline']
                    if 'ExecutionStage' in act_df.columns:
                        sort_cols_act.append('ExecutionStage')
                    act_df = act_df.sort_values(by=sort_cols_act, ascending=True, na_position='last').reset_index(drop=True)
                
                # ActivityExecutionOrder: sort by Pipeline A-Z, then by FromExecutionStage smallest to largest
                if not order_df.empty:
                    sort_cols = []
                    if 'Pipeline' in order_df.columns:
                        sort_cols.append('Pipeline')
                    if 'FromExecutionStage' in order_df.columns:
                        sort_cols.append('FromExecutionStage')
                    if sort_cols:
                        order_df = order_df.sort_values(by=sort_cols, ascending=True, na_position='last').reset_index(drop=True)

                # Only replace if we have dataframes; this guarantees all keys become columns on-disk
                if not act_df.empty or not order_df.empty:
                    try:
                        with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as w2:
                            if not act_df.empty:
                                act_df.to_excel(w2, sheet_name='Activities', index=False)
                            if not order_df.empty:
                                order_df.to_excel(w2, sheet_name='ActivityExecutionOrder', index=False)
                        
                        # Apply formatting to match other sheets (header style, auto-filter, freeze panes)
                        try:
                            from openpyxl.utils import get_column_letter
                            from openpyxl.styles import Font, PatternFill
                            
                            wb_tbl = load_workbook(excel_file)
                            for sheet_name in ['Activities', 'ActivityExecutionOrder']:
                                if sheet_name in wb_tbl.sheetnames:
                                    ws_tbl = wb_tbl[sheet_name]
                                    if ws_tbl.max_row > 1:
                                        # Bold headers with gray fill (same as _format_sheet)
                                        for cell in ws_tbl[1]:
                                            cell.font = Font(bold=True)
                                            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                                        # Auto-filter
                                        ws_tbl.auto_filter.ref = f"A1:{get_column_letter(ws_tbl.max_column)}{ws_tbl.max_row}"
                                        # Freeze panes
                                        ws_tbl.freeze_panes = 'A2'
                            
                            # Apply column hiding from config (after pandas rewrite)
                            try:
                                import json
                                config_path = getattr(self, '_enhancement_config_path', 'enhancement_config.json')
                                with open(config_path, 'r') as f:
                                    hide_cfg = json.load(f).get('hide_config', {})
                                if hide_cfg.get('enabled', False):
                                    hidden_columns = hide_cfg.get('hidden_columns', {})
                                    for sheet_name, cols_to_hide in hidden_columns.items():
                                        if sheet_name in wb_tbl.sheetnames and cols_to_hide:
                                            ws = wb_tbl[sheet_name]
                                            headers = {cell.value: cell.column for cell in ws[1] if cell.value}
                                            for col_name in cols_to_hide:
                                                if col_name in headers:
                                                    col_letter = get_column_letter(headers[col_name])
                                                    ws.column_dimensions[col_letter].hidden = True
                                                    ws.column_dimensions[col_letter].width = 0
                            except Exception:
                                pass
                            
                            wb_tbl.save(excel_file)
                        except Exception:
                            pass
                    except Exception:
                        # Fallback: append missing headers and populate values row-wise from self.lookup
                        wb = load_workbook(excel_file)
                        headers_changed = False
                        if 'Activities' in wb.sheetnames:
                            ws = wb['Activities']
                            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                            for col in ('ParseSequence','ExecutionStage','HasDependsOn','DependsOnCount','CycleFlag'):
                                if col not in headers:
                                    ws.cell(row=1, column=len(headers) + 1, value=col)
                                    headers.append(col)
                                    headers_changed = True
                            # populate activity-level values using lookup mapping if available
                            if hasattr(self, 'lookup') and isinstance(self.lookup.get('activities', None), dict):
                                lookup = self.lookup.get('activities')
                                header_index = {h: i+1 for i, h in enumerate(headers)}
                                max_row = ws.max_row
                                for r in range(2, max_row+1):
                                    try:
                                        p = ws.cell(row=r, column=1).value
                                        a = ws.cell(row=r, column=5).value
                                        key = (p, a)
                                        info = lookup.get(key, {})
                                        if 'ParseSequence' in header_index:
                                            ws.cell(row=r, column=header_index['ParseSequence'], value=info.get('Sequence') if info else '')
                                        if 'ExecutionStage' in header_index:
                                            ws.cell(row=r, column=header_index['ExecutionStage'], value=info.get('ExecutionStage') if info else '')
                                        if 'HasDependsOn' in header_index:
                                            ws.cell(row=r, column=header_index['HasDependsOn'], value='Yes' if info and info.get('Dependencies') else 'No')
                                        if 'DependsOnCount' in header_index:
                                            deps = info.get('Dependencies') if info else []
                                            ws.cell(row=r, column=header_index['DependsOnCount'], value=len(deps) if deps else 0)
                                        if 'CycleFlag' in header_index:
                                            ws.cell(row=r, column=header_index['CycleFlag'], value='Yes' if info and info.get('ExecutionStage') == 'CYCLE' else 'No')
                                    except Exception:
                                        continue
                        if 'ActivityExecutionOrder' in wb.sheetnames:
                            ws2 = wb['ActivityExecutionOrder']
                            headers2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
                            if 'FromExecutionStage' not in headers2:
                                ws2.cell(row=1, column=len(headers2) + 1, value='FromExecutionStage')
                                headers2.append('FromExecutionStage')
                            if 'ToExecutionStage' not in headers2:
                                ws2.cell(row=1, column=len(headers2) + 1, value='ToExecutionStage')
                                headers2.append('ToExecutionStage')
                            if hasattr(self, 'lookup') and isinstance(self.lookup.get('activities', None), dict):
                                lookup = self.lookup.get('activities')
                                header_index2 = {h: i+1 for i, h in enumerate(headers2)}
                                max_row2 = ws2.max_row
                                for r in range(2, max_row2+1):
                                    try:
                                        p = ws2.cell(row=r, column=1).value
                                        from_act = ws2.cell(row=r, column=2).value
                                        to_act = ws2.cell(row=r, column=5).value
                                        from_stage = lookup.get((p, from_act), {}).get('ExecutionStage', '')
                                        to_stage = lookup.get((p, to_act), {}).get('ExecutionStage', '')
                                        if 'FromExecutionStage' in header_index2:
                                            ws2.cell(row=r, column=header_index2['FromExecutionStage'], value=from_stage)
                                        if 'ToExecutionStage' in header_index2:
                                            ws2.cell(row=r, column=header_index2['ToExecutionStage'], value=to_stage)
                                    except Exception:
                                        continue
                        # Save workbook if we changed headers or populated values
                        try:
                            if headers_changed:
                                wb.save(excel_file)
                                self.logger.info("  ‚úì Injected missing ExecutionStage headers/values into workbook")
                        except Exception:
                            pass
                        
                        # Apply formatting to match other sheets (header style, auto-filter, freeze panes)
                        try:
                            from openpyxl import load_workbook
                            from openpyxl.utils import get_column_letter
                            from openpyxl.styles import Font, PatternFill
                            
                            wb2 = load_workbook(excel_file)
                            for sheet_name in ['Activities', 'ActivityExecutionOrder']:
                                if sheet_name in wb2.sheetnames:
                                    ws = wb2[sheet_name]
                                    if ws.max_row > 1:
                                        # Bold headers with gray fill
                                        for cell in ws[1]:
                                            cell.font = Font(bold=True)
                                            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                                        # Auto-filter
                                        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                                        # Freeze panes
                                        ws.freeze_panes = 'A2'
                            wb2.save(excel_file)
                        except Exception:
                            pass
            except Exception:
                pass

            shutil.copy(excel_file, archive_file)
            self.logger.info(f" Archive saved: {archive_file}")

            self._auto_copy_to_streamlit(excel_file)

        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            traceback.print_exc()
            raise

        # CSV sidecar exports removed: all metadata is injected into the Excel workbook.

    analyzer_class.export_to_excel = enhanced_export_to_excel

    print("   Enhanced export_to_excel() applied")

def create_enhanced_beautification_method(analyzer_class):
    """
     CREATE ENHANCED BEAUTIFICATION METHOD

    Replaces _apply_enterprise_formatting() with enhanced version
    """

    def _apply_enhanced_beautification(self, writer):
        """
         APPLY COMPLETE ENHANCED BEAUTIFICATION

        Phase 1: Basic Formatting (all sheets)
        Phase 2: Conditional Formatting (all sheets)
        Phase 3: Hyperlinks (Summary sheet)
        Phase 4: Sheet Protection (optional)
        Phase 5: Page Setup (all sheets)
        """

        workbook = writer.book
        all_sheet_names = [ws.title for ws in workbook.worksheets]

        # Load runtime enhancement config (respects dashboard toggle granularity)
        try:
            _cfg = ENHANCEMENT_CONFIG.get('excel_enhancements', {})
        except Exception:
            _cfg = {}

        core_cfg = _cfg.get('core_formatting', {}) if isinstance(_cfg.get('core_formatting'), dict) else {}
        cond_cfg = _cfg.get('conditional_formatting', {}) if isinstance(_cfg.get('conditional_formatting'), dict) else {}
        link_cfg = _cfg.get('hyperlinks', {}) if isinstance(_cfg.get('hyperlinks'), dict) else {}
        prot_cfg = _cfg.get('protection', {}) if isinstance(_cfg.get('protection'), dict) else {}
        page_cfg = _cfg.get('page_setup', {}) if isinstance(_cfg.get('page_setup'), dict) else {}

        # PHASE 1: Core formatting
        if core_cfg.get('enabled', True):
            self.logger.info("  Phase 1/5: Basic formatting (columns, borders, alignment)...")
            for worksheet in workbook.worksheets:
                try:
                    MasterFormatter.format_worksheet(
                        worksheet,
                        sheet_name=worksheet.title,
                        header_row=1,
                        enable_features={
                            'column_sizing': core_cfg.get('column_sizing', True),
                            'number_format': core_cfg.get('number_format', True),
                            'alignment': core_cfg.get('alignment', True),
                            'borders': core_cfg.get('borders', True),
                            'row_shading': core_cfg.get('row_shading', True),
                            'header_style': core_cfg.get('header_style', True)
                        }
                    )
                except Exception as e:
                    self.logger.warning(f"Basic formatting failed for {worksheet.title}: {e}")
        else:
            self.logger.info("  Phase 1/5: Basic formatting skipped (disabled)")

        # PHASE 2: Conditional formatting
        if cond_cfg.get('enabled', True):
            self.logger.info("  Phase 2/5: Conditional formatting (data bars, icons, colors)...")
            for worksheet in workbook.worksheets:
                try:
                    MasterConditionalFormatter.apply_all_conditional_formatting(
                        worksheet,
                        sheet_name=worksheet.title,
                        header_row=1,
                        enable_features={
                            'data_bars': cond_cfg.get('data_bars', True),
                            'icon_sets': cond_cfg.get('icon_sets', True),
                            'color_scales': cond_cfg.get('color_scales', True),
                            'status_highlighting': cond_cfg.get('status_highlighting', True)
                        }
                    )
                except Exception as e:
                    self.logger.warning(f"Conditional formatting failed for {worksheet.title}: {e}")
        else:
            self.logger.info("  Phase 2/5: Conditional formatting skipped (disabled)")

        for worksheet in workbook.worksheets:
            sheet_name = worksheet.title.lower()

            try:
                if sheet_name == 'summary':
                    SpecialSheetFormatters.format_summary_sheet(worksheet)
                elif 'impact' in sheet_name:
                    SpecialSheetFormatters.format_impact_analysis_sheet(worksheet)
                elif 'circular' in sheet_name:
                    SpecialSheetFormatters.format_circular_dependencies_sheet(worksheet)
            except Exception as e:
                self.logger.warning(f"Special formatting failed for {worksheet.title}: {e}")

        # PHASE 3: Hyperlinks
        if link_cfg.get('enabled', True):
            self.logger.info("  Phase 3/5: Adding hyperlinks and navigation...")
            
            # Skip table formatting - use plain filters instead for minimal white appearance
            # Apply auto-filter to all data sheets for plain appearance
            for worksheet in workbook.worksheets:
                sheet_name = worksheet.title
                # Skip Summary sheet (has custom layout)
                if 'Summary' in sheet_name:
                    continue
                
                # Apply auto-filter only (no table formatting for plain white look)
                if worksheet.max_row > 1:
                    try:
                        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                    except Exception:
                        pass
            
            if 'Summary' in workbook.sheetnames:
                summary_ws = workbook['Summary']
                try:
                    if link_cfg.get('auto_convert_references', True):
                        HyperlinkManager.auto_convert_sheet_references(
                            summary_ws,
                            available_sheets=all_sheet_names,
                            header_row=1
                        )
                    if link_cfg.get('summary_navigation', True):
                        HyperlinkManager.add_navigation_links_to_summary(
                            summary_ws,
                            all_sheets=all_sheet_names
                        )
                    self.logger.info("   Hyperlinks added to Summary sheet")
                except Exception as e:
                    self.logger.warning(f"Hyperlink creation failed: {e}")
        else:
            self.logger.info("  Phase 3/5: Hyperlinks skipped (disabled)")

        for worksheet in workbook.worksheets:
            try:
                CellCommentManager.auto_add_helpful_comments(
                    worksheet,
                    sheet_name=worksheet.title,
                    header_row=1
                )
            except Exception as e:
                pass  # Comments are optional

        # PHASE 4: Sheet Protection
        if prot_cfg.get('enabled', False):
            self.logger.info("  Phase 4/5: Applying sheet protection...")
            try:
                sheets_to_protect = [
                    ws.title for ws in workbook.worksheets
                    if ws.title.lower() != 'summary'
                ]
                SheetProtectionManager.protect_all_sheets(
                    workbook,
                    sheets_to_protect=sheets_to_protect,
                    password=prot_cfg.get('password') or None
                )
                self.logger.info(f"   Protected {len(sheets_to_protect)} sheets")
            except Exception as e:
                self.logger.warning(f"Sheet protection failed: {e}")
        else:
            self.logger.info("  Phase 4/5: Sheet protection skipped (disabled)")

        # PHASE 5: Page setup
        if page_cfg.get('enabled', True):
            self.logger.info("  Phase 5/5: Configuring print settings...")
            try:
                PageSetupManager.auto_setup_all_sheets(workbook)
                self.logger.info("   Print settings configured")
            except Exception as e:
                self.logger.warning(f"Page setup failed: {e}")
        else:
            self.logger.info("  Phase 5/5: Page setup skipped (disabled)")

        # PHASE 6: Hide sheets and columns (config-based)
        hide_cfg = ENHANCEMENT_CONFIG.get('hide_config', {})
        if hide_cfg.get('enabled', False):
            self.logger.info("  Phase 6: Applying hide configuration...")
            
            # Hide specified sheets
            hidden_sheets = hide_cfg.get('hidden_sheets', [])
            for sheet_name in hidden_sheets:
                if sheet_name in workbook.sheetnames:
                    try:
                        workbook[sheet_name].sheet_state = 'hidden'
                        self.logger.info(f"    Hidden sheet: {sheet_name}")
                    except Exception as e:
                        self.logger.warning(f"Failed to hide sheet {sheet_name}: {e}")
            
            # Hide specified columns per sheet
            hidden_columns = hide_cfg.get('hidden_columns', {})
            for sheet_name, columns_to_hide in hidden_columns.items():
                if sheet_name in workbook.sheetnames and columns_to_hide:
                    ws = workbook[sheet_name]
                    try:
                        # Get header row to find column positions
                        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
                        for col_name in columns_to_hide:
                            if col_name in headers:
                                col_letter = get_column_letter(headers[col_name])
                                ws.column_dimensions[col_letter].hidden = True
                                ws.column_dimensions[col_letter].width = 0
                                self.logger.info(f"    Hidden column: {sheet_name}.{col_name}")
                    except Exception as e:
                        self.logger.warning(f"Failed to hide columns in {sheet_name}: {e}")

        self.logger.info(" Enhanced beautification complete!")

    analyzer_class._apply_enhanced_beautification = _apply_enhanced_beautification

    print("   Enhanced beautification method applied")

def apply_excel_enhancements(analyzer_class=None, verbose: bool = True):
    """
     MASTER FUNCTION: Apply ALL Excel enhancements

    This is the main entry point that applies all patches from Parts 1-4.

    Usage:
        Method 1 - Explicit class:
            from adf_analyzer_v10_complete import UltimateEnterpriseADFAnalyzer
            from adf_analyzer_v10_excel_enhancements import apply_excel_enhancements

            apply_excel_enhancements(UltimateEnterpriseADFAnalyzer)

        Method 2 - Auto-import:
            from adf_analyzer_v10_excel_enhancements import apply_excel_enhancements

            apply_excel_enhancements()  # Auto-imports and patches

    Args:
        analyzer_class: Analyzer class to patch (auto-imports if None)
        verbose: Print progress messages

    Returns:
        True if successful, False otherwise
    """

    if analyzer_class is None:
        try:
            from adf_analyzer_v10_complete import UltimateEnterpriseADFAnalyzer
            analyzer_class = UltimateEnterpriseADFAnalyzer
        except ImportError:
            print(" ERROR: Could not import UltimateEnterpriseADFAnalyzer")
            print("   Make sure adf_analyzer_v10_complete.py is in the same directory")
            return False

    if verbose:
        print("\n" + "="*80)
        print(" APPLYING EXCEL BEAUTIFICATION ENHANCEMENTS")
        print("="*80 + "\n")

    try:

        if verbose:
            print("üì¶ Part 1/4: Core Enhancement Framework...")

        if verbose:
            print("üì¶ Part 2/4: Conditional Formatting...")

        if verbose:
            print("üì¶ Part 3/4: Hyperlinks & Advanced Features...")

        if verbose:
            print("üì¶ Part 4/4: Master Integration...")

        create_enhanced_export_function(analyzer_class)
        create_enhanced_beautification_method(analyzer_class)

        if verbose:
            print("\n" + "="*80)
            print(" EXCEL ENHANCEMENTS APPLIED SUCCESSFULLY")
            print("="*80)
            print("\n New Features Added:")
            print("   Intelligent column sizing (content-aware)")
            print("   Professional cell borders & styling")
            print("   Alternating row colors")
            print("   Advanced number formatting (%, thousand separators)")
            print("   Smart text alignment & wrapping")
            print("   Data bars (visual progress indicators)")
            print("   Icon sets (traffic lights, arrows)")
            print("   Color scales (heat maps)")
            print("   Status-based highlighting (CRITICAL=red, etc.)")
            print("    CLICKABLE HYPERLINKS in Summary sheet (CRITICAL FIX!)")
            print("   Navigation section with all sheets")
            print("   Sheet protection (allow filtering)")
            print("   Cell comments/tooltips")
            print("   Professional print settings")
            print("="*80 + "\n")

        return True

    except Exception as e:
        if verbose:
            print(f"\n ENHANCEMENT APPLICATION FAILED: {e}")
            traceback.print_exc()
        return False

class EnhancementValidator:
    """
     ENHANCEMENT VALIDATOR

    Validates that all enhancements are working correctly
    """

    @staticmethod
    def validate_enhancements(excel_file: Path) -> Dict[str, bool]:
        """
         Validate Excel file has all enhancements

        Args:
            excel_file: Path to Excel file

        Returns:
            Dict of validation results
        """

        from openpyxl import load_workbook

        results = {
            'file_exists': False,
            'has_multiple_sheets': False,
            'has_summary_sheet': False,
            'has_hyperlinks': False,
            'has_conditional_formatting': False,
            'has_formatted_headers': False,
            'has_borders': False,
            'columns_sized': False,
            'has_protection': False
        }

        try:

            if not excel_file.exists():
                return results

            results['file_exists'] = True

            wb = load_workbook(excel_file)

            if len(wb.sheetnames) > 1:
                results['has_multiple_sheets'] = True

            if 'Summary' in wb.sheetnames:
                results['has_summary_sheet'] = True

                summary_ws = wb['Summary']

                for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row):
                    for cell in row:
                        if cell.hyperlink:
                            results['has_hyperlinks'] = True
                            break
                    if results['has_hyperlinks']:
                        break

                first_row = summary_ws[1]
                for cell in first_row:
                    if cell.font and cell.font.bold:
                        results['has_formatted_headers'] = True
                        break

                if summary_ws.max_row > 1:
                    test_cell = summary_ws.cell(2, 1)
                    if test_cell.border and test_cell.border.left:
                        results['has_borders'] = True

                col_width = summary_ws.column_dimensions['A'].width
                if col_width and col_width > 8:
                    results['columns_sized'] = True

            for ws in wb.worksheets:
                if ws.conditional_formatting:
                    results['has_conditional_formatting'] = True
                    break

            for ws in wb.worksheets:
                if ws.protection.sheet:
                    results['has_protection'] = True
                    break

            wb.close()

        except Exception as e:
            print(f"  Validation error: {e}")

        return results

    @staticmethod
    def print_validation_report(results: Dict[str, bool]):
        """Print validation report"""

        print("\n" + "="*80)
        print(" ENHANCEMENT VALIDATION REPORT")
        print("="*80 + "\n")

        checks = [
            ('file_exists', 'File exists'),
            ('has_multiple_sheets', 'Multiple sheets created'),
            ('has_summary_sheet', 'Summary sheet exists'),
            ('has_hyperlinks', ' Hyperlinks present (CRITICAL FIX)'),
            ('has_conditional_formatting', 'Conditional formatting applied'),
            ('has_formatted_headers', 'Headers formatted'),
            ('has_borders', 'Cell borders applied'),
            ('columns_sized', 'Columns auto-sized'),
            ('has_protection', 'Sheet protection enabled')
        ]

        passed = 0
        total = len(checks)

        for key, description in checks:
            status = " PASS" if results.get(key, False) else " FAIL"
            print(f"  {status}  {description}")
            if results.get(key, False):
                passed += 1

        print("\n" + "-"*80)
        print(f"Score: {passed}/{total} checks passed ({passed/total*100:.0f}%)")
        print("="*80 + "\n")

        if passed == total:
            print("üéâ ALL ENHANCEMENTS VALIDATED SUCCESSFULLY!")
        elif passed >= total * 0.7:
            print("  Most enhancements working, some issues detected")
        else:
            print(" Multiple enhancement failures detected")

        return passed == total

def print_usage_guide():
    """Print complete usage guide"""
    
    print("""
Usage Guide:
1. Import the enhancement module
2. Apply enhancements to your workbook
3. Save the enhanced Excel file
    """)

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_enhanced_summary_sheet_writer(analyzer_class):
    """
     REPLACE ORIGINAL _write_summary_sheet WITH ENHANCED VERSION

    Creates beautiful summary with:
    - Professional banner
    - Executive summary
    - Key metrics dashboard
    - Critical alerts
    - Recommendations
    - Navigation links
    """

    def _write_enhanced_summary_sheet(self, writer, timestamp: str):
        """
         ENHANCED SUMMARY SHEET

        Layout:
        1. Project Banner (rows 1-6)
        2. Executive Summary (rows 8-12)
        3. Critical Alerts (rows 14+)
        4. Key Metrics Dashboard (rows 20+)
        5. Resource Overview (rows 35+)
        6. Navigation Links (bottom)
        """

        import pandas as pd

        workbook = writer.book

        df_init = pd.DataFrame({'_': ['']})
        sheet_name = self._get_unique_sheet_name('Summary')
        df_init.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        ws = writer.sheets[sheet_name]

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50

        current_row = 1

        current_row = self._write_project_banner(ws, current_row, timestamp)

        current_row += 2
        current_row = self._write_executive_summary(ws, current_row)

        current_row += 2
        current_row = self._write_critical_alerts(ws, current_row)

        current_row += 2
        current_row = self._write_metrics_dashboard(ws, current_row)

        current_row += 2
        current_row = self._write_resource_overview(ws, current_row)

        current_row += 2
        current_row = self._write_recommendations(ws, current_row)

        current_row += 2
        current_row = self._write_detailed_statistics(ws, current_row, timestamp)

        self.logger.info(f"  ‚úì Enhanced Summary")

    def _write_project_banner(self, ws, start_row: int, timestamp: str) -> int:
        """
         Write beautiful project banner

        ‚îå‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îê
        ‚îÇ  üè≠ AZURE DATA FACTORY - ARM TEMPLATE ANALYSIS REPORT            ‚îÇ
        ‚îÇ  Enterprise-Grade Architecture Assessment                        ‚îÇ
        ‚îÇ  Generated: 2024-01-15 14:30:45                                 ‚îÇ
        ‚îî‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îò
        """

        title_cell = ws.cell(start_row, 1)
        title_cell.value = "üè≠ AZURE DATA FACTORY - ARM TEMPLATE ANALYSIS REPORT"

        ws.merge_cells(f'A{start_row}:D{start_row}')

        title_cell.font = Font(
            name='Calibri',
            size=18,
            bold=True,
            color='FFFFFF'
        )
        title_cell.fill = PatternFill(
            start_color='0066CC',
            end_color='0066CC',
            fill_type='solid'
        )
        title_cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        start_row += 1
        subtitle_cell = ws.cell(start_row, 1)
        subtitle_cell.value = "Enterprise-Grade Architecture Assessment & Comprehensive Analysis"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        subtitle_cell.font = Font(
            name='Calibri',
            size=12,
            italic=True,
            color='FFFFFF'
        )
        subtitle_cell.fill = PatternFill(
            start_color='0099FF',
            end_color='0099FF',
            fill_type='solid'
        )
        subtitle_cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        start_row += 1

        ws.cell(start_row, 1).value = "üìÑ Source Template:"
        ws.cell(start_row, 2).value = str(self.json_path)
        ws.cell(start_row, 1).font = Font(bold=True)

        start_row += 1

        ws.cell(start_row, 1).value = "üìÖ Analysis Date:"
        ws.cell(start_row, 2).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(start_row, 1).font = Font(bold=True)

        start_row += 1

        ws.cell(start_row, 1).value = "üîß Analyzer Version:"
        ws.cell(start_row, 2).value = "v10.0 - Production Ready (Enhanced Edition)"
        ws.cell(start_row, 1).font = Font(bold=True)

        start_row += 1

        ws.cell(start_row, 1).value = "üë§ Generated By:"
        ws.cell(start_row, 2).value = "Ultimate Enterprise ADF Analyzer"
        ws.cell(start_row, 1).font = Font(bold=True)

        return start_row + 1

    def _write_executive_summary(self, ws, start_row: int) -> int:
        """
         Write executive summary section
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " EXECUTIVE SUMMARY"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(
            name='Calibri',
            size=14,
            bold=True,
            color='FFFFFF'
        )
        header_cell.fill = PatternFill(
            start_color='2F5496',
            end_color='2F5496',
            fill_type='solid'
        )
        header_cell.alignment = Alignment(
            horizontal='left',
            vertical='center'
        )

        start_row += 1

        total_resources = len(self.resources['all'])
        total_pipelines = len(self.resources['pipelines'])
        total_activities = len(self.results['activities'])
        total_dataflows = len(self.resources['dataflows'])

        orphaned_count = (
            len(self.results['orphaned_pipelines']) +
            len(self.results['orphaned_dataflows']) +
            len(self.results['orphaned_datasets']) +
            len(self.results['orphaned_linked_services'])
        )

        circular_deps = len(self.results['circular_dependencies'])

        summary_items = [
            (" Total Resources Analyzed", total_resources, "All ARM template resources"),
            ("üîÑ Active Pipelines", total_pipelines, "Data orchestration workflows"),
            ("‚ö° Total Activities", total_activities, "Execution steps across all pipelines"),
            ("üåä Data Flows", total_dataflows, "ETL transformation flows"),
        ]

        for label, value, description in summary_items:
            ws.cell(start_row, 1).value = label
            ws.cell(start_row, 2).value = value
            ws.cell(start_row, 3).value = description

            ws.cell(start_row, 1).font = Font(bold=True, size=11)
            ws.cell(start_row, 2).font = Font(size=11, bold=True, color='0066CC')
            ws.cell(start_row, 3).font = Font(size=10, italic=True)

            start_row += 1

        return start_row

    def _write_critical_alerts(self, ws, start_row: int) -> int:
        """
         Write critical alerts section
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üö® CRITICAL ALERTS & ACTION ITEMS"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(
            name='Calibri',
            size=14,
            bold=True,
            color='FFFFFF'
        )
        header_cell.fill = PatternFill(
            start_color='C00000',
            end_color='C00000',
            fill_type='solid'
        )
        header_cell.alignment = Alignment(
            horizontal='left',
            vertical='center'
        )

        start_row += 1

        circular_deps = len(self.results['circular_dependencies'])
        orphaned_pipelines = len(self.results['orphaned_pipelines'])
        broken_triggers = len([t for t in self.results['orphaned_triggers'] if t.get('Type') == 'BrokenReference'])

        critical_impact_pipelines = len([
            p for p in self.results['impact_analysis']
            if p.get('Impact') == 'CRITICAL'
        ])

        alerts = []

        if circular_deps > 0:
            alerts.append({
                'icon': '',
                'severity': 'CRITICAL',
                'issue': f'{circular_deps} Circular Dependencies Detected',
                'action': 'Fix immediately - can cause infinite loops',
                'sheet': 'CircularDependencies'
            })

        if broken_triggers > 0:
            alerts.append({
                'icon': '',
                'severity': 'HIGH',
                'issue': f'{broken_triggers} Broken Trigger References',
                'action': 'Update trigger pipeline references',
                'sheet': 'OrphanedTriggers'
            })

        if orphaned_pipelines > 10:
            alerts.append({
                'icon': '',
                'severity': 'MEDIUM',
                'issue': f'{orphaned_pipelines} Orphaned Pipelines',
                'action': 'Review and clean up unused pipelines',
                'sheet': 'OrphanedPipelines'
            })

        if critical_impact_pipelines > 0:
            alerts.append({
                'icon': '‚Ñπ',
                'severity': 'INFO',
                'issue': f'{critical_impact_pipelines} High-Impact Pipelines',
                'action': 'Review dependencies carefully before changes',
                'sheet': 'ImpactAnalysis'
            })

        if alerts:
            for alert in alerts:

                issue_cell = ws.cell(start_row, 1)
                issue_cell.value = f"{alert['icon']} {alert['issue']}"
                issue_cell.font = Font(bold=True, size=11)

                severity_cell = ws.cell(start_row, 2)
                severity_cell.value = alert['severity']
                severity_cell.font = Font(bold=True)

                if alert['severity'] == 'CRITICAL':
                    severity_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    severity_cell.font = Font(bold=True, color='FFFFFF')
                elif alert['severity'] == 'HIGH':
                    severity_cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                    severity_cell.font = Font(bold=True, color='FFFFFF')
                elif alert['severity'] == 'MEDIUM':
                    severity_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                ws.cell(start_row, 3).value = alert['action']
                ws.cell(start_row, 3).font = Font(size=10)

                ws.cell(start_row, 4).value = f" See {alert['sheet']}"
                ws.cell(start_row, 4).font = Font(size=10, color='0563C1', underline='single')

                start_row += 1
        else:

            ws.cell(start_row, 1).value = " No critical issues detected"
            ws.cell(start_row, 1).font = Font(bold=True, color='00B050', size=11)
            ws.merge_cells(f'A{start_row}:D{start_row}')
            start_row += 1

        return start_row

    def _write_metrics_dashboard(self, ws, start_row: int) -> int:
        """
         Write key metrics dashboard
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " KEY METRICS DASHBOARD"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 2

        metrics = [

            [
                ("Total Pipelines", len(self.resources['pipelines']), "üîÑ"),
                ("Total Activities", len(self.results['activities']), "‚ö°"),
                ("Data Flows", len(self.resources['dataflows']), "üåä")
            ],

            [
                ("Datasets", len(self.resources['datasets']), ""),
                ("Linked Services", len(self.resources['linkedServices']), "üîó"),
                ("Triggers", len(self.resources['triggers']), "‚è∞")
            ]
        ]

        col_offset = 0
        for row_metrics in metrics:
            col = 1
            for label, value, icon in row_metrics:

                metric_cell = ws.cell(start_row, col)
                metric_cell.value = f"{icon} {label}"
                metric_cell.font = Font(bold=True, size=10)
                metric_cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
                metric_cell.alignment = Alignment(horizontal='center', vertical='center')
                metric_cell.border = ExcelBorders.thin_border()

                value_cell = ws.cell(start_row + 1, col)
                value_cell.value = value
                value_cell.font = Font(size=16, bold=True, color='0066CC')
                value_cell.alignment = Alignment(horizontal='center', vertical='center')
                value_cell.border = ExcelBorders.thin_border()

                col += 1

            start_row += 2

        return start_row + 1

    def _write_resource_overview(self, ws, start_row: int) -> int:
        """
         Write resource overview section
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üì¶ RESOURCE OVERVIEW"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        headers = ['Category', 'Resource Type', 'Count', 'Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(start_row, col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = ExcelBorders.thin_border()

        start_row += 1

        resources_data = [
            ('CORE RESOURCES', 'Pipelines', len(self.resources['pipelines']), ' PipelineAnalysis'),
            ('', 'DataFlows', len(self.resources['dataflows']), ' DataFlows'),
            ('', 'Datasets', len(self.resources['datasets']), ' Datasets'),
            ('', 'Linked Services', len(self.resources['linkedServices']), ' LinkedServices'),
            ('', 'Triggers', len(self.resources['triggers']), ' Triggers'),
            ('', 'Integration Runtimes', len(self.resources['integrationRuntimes']), ' IntegrationRuntimes'),
            ('ANALYSIS', 'Activity Dependencies', len(self.results['activity_execution_order']), ' ActivityExecutionOrder'),
            ('', 'Data Lineage Records', len(self.results['data_lineage']), ' DataLineage'),
            ('', 'Circular Dependencies', len(self.results['circular_dependencies']), ' CircularDependencies'),
            ('QUALITY', 'Orphaned Pipelines', len(self.results['orphaned_pipelines']), ' OrphanedPipelines'),
            ('', 'Orphaned Datasets', len(self.results['orphaned_datasets']), ' OrphanedDatasets'),
        ]

        for category, resource_type, count, link in resources_data:
            ws.cell(start_row, 1).value = category
            ws.cell(start_row, 2).value = resource_type
            ws.cell(start_row, 3).value = count
            ws.cell(start_row, 4).value = link

            if category:
                ws.cell(start_row, 1).font = Font(bold=True)

            ws.cell(start_row, 3).font = Font(bold=True, color='0066CC')
            ws.cell(start_row, 4).font = Font(color='0563C1', underline='single')

            for col in range(1, 5):
                ws.cell(start_row, col).border = ExcelBorders.thin_border()

            start_row += 1

        return start_row

    def _write_recommendations(self, ws, start_row: int) -> int:
        """
         Write recommendations section
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " RECOMMENDATIONS & NEXT STEPS"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        recommendations = []

        if self.results['circular_dependencies']:
            recommendations.append(
                "1.  URGENT: Fix circular dependencies immediately - they can cause infinite execution loops"
            )

        total_orphaned = (
            len(self.results['orphaned_pipelines']) +
            len(self.results['orphaned_datasets']) +
            len(self.results['orphaned_linked_services'])
        )

        if total_orphaned > 20:
            recommendations.append(
                f"2. üßπ Clean up {total_orphaned} orphaned resources to reduce maintenance overhead"
            )

        critical_pipelines = [
            p for p in self.results['impact_analysis']
            if p.get('Impact') == 'CRITICAL'
        ]

        if critical_pipelines:
            recommendations.append(
                f"3.   Review {len(critical_pipelines)} critical-impact pipelines before making changes"
            )

        stopped_triggers = [
            t for t in self.results['triggers']
            if t.get('State') == 'Stopped'
        ]

        if stopped_triggers:
            recommendations.append(
                f"4. ‚è∏  Investigate {len(stopped_triggers)} stopped triggers - are they intentional?"
            )

        recommendations.append(
            "5.  Use ImpactAnalysis sheet to understand dependencies before modifications"
        )

        recommendations.append(
            "6.  Review DataLineage sheet for end-to-end data flow understanding"
        )

        recommendations.append(
            "7.  Monitor activity counts for overly complex pipelines (>50 activities)"
        )

        for rec in recommendations:
            ws.cell(start_row, 1).value = rec
            ws.merge_cells(f'A{start_row}:D{start_row}')
            ws.cell(start_row, 1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws.cell(start_row, 1).font = Font(size=10)

            if '' in rec:
                ws.cell(start_row, 1).fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            elif '' in rec:
                ws.cell(start_row, 1).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

            ws.row_dimensions[start_row].height = 25
            start_row += 1

        return start_row

    def _write_detailed_statistics(self, ws, start_row: int, timestamp: str) -> int:
        """
         Write detailed statistics (original content)
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " DETAILED STATISTICS"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        return start_row

    analyzer_class._write_summary_sheet = _write_enhanced_summary_sheet
    analyzer_class._write_project_banner = _write_project_banner
    analyzer_class._write_executive_summary = _write_executive_summary
    analyzer_class._write_critical_alerts = _write_critical_alerts
    analyzer_class._write_metrics_dashboard = _write_metrics_dashboard
    analyzer_class._write_resource_overview = _write_resource_overview
    analyzer_class._write_recommendations = _write_recommendations
    analyzer_class._write_detailed_statistics = _write_detailed_statistics

    print("   Enhanced Summary Sheet writer applied")

def apply_excel_enhancements_with_summary(analyzer_class=None, verbose: bool = True):
    """
     ENHANCED VERSION: Apply ALL Excel enhancements INCLUDING beautiful summary
    """

    if analyzer_class is None:
        try:
            from adf_analyzer_v10_complete import UltimateEnterpriseADFAnalyzer
            analyzer_class = UltimateEnterpriseADFAnalyzer
        except ImportError:
            print(" ERROR: Could not import UltimateEnterpriseADFAnalyzer")
            return False

    if verbose:
        print("\n" + "="*80)
        print(" APPLYING EXCEL BEAUTIFICATION ENHANCEMENTS (WITH ENHANCED SUMMARY)")
        print("="*80 + "\n")

    try:

        if verbose:
            print("üì¶ Parts 1-3: Core formatting, conditional formatting, hyperlinks...")

        if verbose:
            print("üì¶ Part 4: Master integration...")

        create_enhanced_export_function(analyzer_class)
        create_enhanced_beautification_method(analyzer_class)

        if verbose:
            print("üì¶ Part 5: Enhanced Summary Sheet...")

        create_enhanced_summary_sheet_writer(analyzer_class)

        if verbose:
            print("\n" + "="*80)
            print(" ALL EXCEL ENHANCEMENTS APPLIED SUCCESSFULLY")
            print("="*80)
            print("\n New Features Added:")
            print("   Beautiful project banner in Summary sheet")
            print("   Executive summary section")
            print("   Critical alerts dashboard")
            print("   Key metrics visualization")
            print("   Automated recommendations")
            print("   Professional formatting throughout")
            print("   Clickable hyperlinks")
            print("   Data bars, icon sets, color scales")
            print("   Sheet protection")
            print("="*80 + "\n")

        return True

    except Exception as e:
        if verbose:
            print(f"\n ENHANCEMENT APPLICATION FAILED: {e}")
            traceback.print_exc()
        return False

print(" Part 5/6 loaded: Enhanced Summary Sheet module loaded")
"""Excel Enhancement Module - Professional Excel formatting and styling capabilities"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter
import math

def add_advanced_summary_sections(analyzer_class):
    """
     ADD ADVANCED SUMMARY SECTIONS

    New sections:
    1. Health Score Dashboard
    2. Cost Analysis & Optimization
    3. Complexity Heat Map
    4. Performance Insights
    5. Top Pipelines Ranking
    6. Security & Compliance
    7. Activity Distribution
    8. Data Flow Network Stats
    9. Change Risk Assessment
    10. Quick Action Buttons
    """

    def _write_health_score_dashboard(self, ws, start_row: int) -> int:
        """
         HEALTH SCORE DASHBOARD

        Visual health indicators:
        - Overall Health Score (0-100)
        - Quality Score
        - Performance Score
        - Security Score
        - Maintainability Score
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üè• FACTORY HEALTH SCORE DASHBOARD"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        quality_score = self._calculate_quality_score()
        performance_score = self._calculate_performance_score()
        security_score = self._calculate_security_score()
        maintainability_score = self._calculate_maintainability_score()

        overall_health = int(
            quality_score * 0.3 +
            performance_score * 0.2 +
            security_score * 0.3 +
            maintainability_score * 0.2
        )

        ws.cell(start_row, 1).value = "OVERALL HEALTH"
        ws.cell(start_row, 1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{start_row}:B{start_row}')

        health_cell = ws.cell(start_row, 3)
        health_cell.value = f"{overall_health}/100"
        health_cell.font = Font(size=24, bold=True, color=self._get_health_color(overall_health))
        health_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'C{start_row}:D{start_row}')

        ws.cell(start_row + 1, 3).value = self._get_health_status(overall_health)
        ws.cell(start_row + 1, 3).font = Font(bold=True, size=11)
        ws.merge_cells(f'C{start_row + 1}:D{start_row + 1}')
        ws.cell(start_row + 1, 3).alignment = Alignment(horizontal='center')

        start_row += 3

        scores = [
            ("Quality Score", quality_score, "Code quality, circular deps, orphaned resources"),
            ("Performance Score", performance_score, "Pipeline efficiency, activity counts"),
            ("Security Score", security_score, "Key Vault usage, IR security, permissions"),
            ("Maintainability Score", maintainability_score, "Complexity, documentation, naming")
        ]

        for label, score, description in scores:

            ws.cell(start_row, 1).value = label
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

            score_cell = ws.cell(start_row, 2)
            score_cell.value = f"{score}/100"
            score_cell.font = Font(size=11, bold=True, color=self._get_health_color(score))
            score_cell.alignment = Alignment(horizontal='center')

            progress_cell = ws.cell(start_row, 3)
            progress_cell.value = "‚ñà" * int(score / 10)
            progress_cell.font = Font(size=14, color=self._get_health_color(score))

            ws.cell(start_row, 4).value = description
            ws.cell(start_row, 4).font = Font(size=9, italic=True)

            start_row += 1

        return start_row + 1

    def _write_cost_analysis(self, ws, start_row: int) -> int:
        """
         COST ANALYSIS & OPTIMIZATION OPPORTUNITIES

        Estimates:
        - DIU hours consumption
        - Pipeline execution frequency
        - Resource optimization opportunities
        - Cost-saving recommendations
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üí∞ COST ANALYSIS & OPTIMIZATION OPPORTUNITIES"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        total_pipelines = len(self.resources['pipelines'])
        total_copy_activities = len([a for a in self.results['activities'] if a.get('ActivityType') == 'Copy'])
        total_dataflows = len(self.resources['dataflows'])

        estimated_diu_hours = total_copy_activities * 2  # Avg 2 DIU hours per copy
        estimated_monthly_cost = estimated_diu_hours * 0.25  # $0.25 per DIU-hour (example)

        ws.cell(start_row, 1).value = "Resource Type"
        ws.cell(start_row, 2).value = "Count"
        ws.cell(start_row, 3).value = "Est. Monthly Cost"
        ws.cell(start_row, 4).value = "Optimization Potential"

        for col in range(1, 5):
            cell = ws.cell(start_row, col)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.border = ExcelBorders.thin_border()

        start_row += 1

        cost_items = [
            ("Copy Activities", total_copy_activities, f"${estimated_monthly_cost:.2f}", "Use staging for large datasets"),
            ("Data Flows", total_dataflows, f"${total_dataflows * 15:.2f}", "Optimize compute settings"),
            ("Pipeline Executions", total_pipelines, "Depends on triggers", "Review trigger schedules"),
            ("Integration Runtimes", len(self.resources['integrationRuntimes']), "Varies", "Consider shared IRs"),
        ]

        for resource_type, count, cost, optimization in cost_items:
            ws.cell(start_row, 1).value = resource_type
            ws.cell(start_row, 2).value = count
            ws.cell(start_row, 2).font = Font(bold=True, color='0066CC')
            ws.cell(start_row, 3).value = cost
            ws.cell(start_row, 4).value = optimization
            ws.cell(start_row, 4).font = Font(size=9, italic=True)

            for col in range(1, 5):
                ws.cell(start_row, col).border = ExcelBorders.thin_border()

            start_row += 1

        start_row += 1
        ws.cell(start_row, 1).value = " Cost Optimization Opportunities:"
        ws.cell(start_row, 1).font = Font(bold=True, size=11, color='FF6600')
        ws.merge_cells(f'A{start_row}:D{start_row}')
        start_row += 1

        orphaned_count = len(self.results['orphaned_pipelines']) + len(self.results['orphaned_datasets'])
        potential_savings = orphaned_count * 5  # $5 per unused resource/month

        opportunities = [
            f"‚Ä¢ Remove {orphaned_count} orphaned resources ‚Üí Save ~${potential_savings}/month",
            f"‚Ä¢ Consolidate {len(self.resources['integrationRuntimes'])} Integration Runtimes if possible",
            f"‚Ä¢ Review trigger schedules for {len(self.resources['triggers'])} triggers",
            "‚Ä¢ Enable staging for large Copy activities to reduce DIU consumption",
            "‚Ä¢ Use incremental loading instead of full loads where applicable"
        ]

        for opp in opportunities:
            ws.cell(start_row, 1).value = opp
            ws.merge_cells(f'A{start_row}:D{start_row}')
            ws.cell(start_row, 1).font = Font(size=10)
            start_row += 1

        return start_row + 1

    def _write_complexity_heat_map(self, ws, start_row: int) -> int:
        """
         COMPLEXITY HEAT MAP

        Visual representation of pipeline complexity distribution
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üå° COMPLEXITY HEAT MAP"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        complexity_distribution = {
            'Critical (100+)': 0,
            'High (50-99)': 0,
            'Medium (20-49)': 0,
            'Low (<20)': 0
        }

        for pipeline in self.results['pipeline_analysis']:
            score = pipeline.get('ComplexityScore', 0)

            if score >= 100:
                complexity_distribution['Critical (100+)'] += 1
            elif score >= 50:
                complexity_distribution['High (50-99)'] += 1
            elif score >= 20:
                complexity_distribution['Medium (20-49)'] += 1
            else:
                complexity_distribution['Low (<20)'] += 1

        total_pipelines = sum(complexity_distribution.values())

        colors = {
            'Critical (100+)': 'C00000',
            'High (50-99)': 'FF6600',
            'Medium (20-49)': 'FFC000',
            'Low (<20)': '92D050'
        }

        for level, count in complexity_distribution.items():
            percentage = (count / total_pipelines * 100) if total_pipelines > 0 else 0

            ws.cell(start_row, 1).value = level
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

            ws.cell(start_row, 2).value = count
            ws.cell(start_row, 2).font = Font(bold=True, size=11)
            ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

            bar_length = int(percentage / 5)  # Scale to fit
            bar_cell = ws.cell(start_row, 3)
            bar_cell.value = "‚ñà" * bar_length
            bar_cell.font = Font(size=14, color=colors[level])

            ws.cell(start_row, 4).value = f"{percentage:.1f}%"
            ws.cell(start_row, 4).font = Font(size=10)
            ws.cell(start_row, 4).fill = PatternFill(
                start_color=colors[level],
                end_color=colors[level],
                fill_type='solid'
            )
            ws.cell(start_row, 4).font = Font(bold=True, color='FFFFFF')
            ws.cell(start_row, 4).alignment = Alignment(horizontal='center')

            start_row += 1

        return start_row + 1

    def _write_dataflow_complexity_heat_map(self, ws, start_row: int) -> int:
        """
         DATAFLOW COMPLEXITY HEAT MAP

        Visual representation of dataflow transformation complexity distribution
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üî¨ DATAFLOW COMPLEXITY HEAT MAP"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=12, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='4B0082', end_color='4B0082', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        complexity_distribution = {
            'Critical (100+)': 0,
            'High (50-99)': 0,
            'Medium (20-49)': 0,
            'Low (<20)': 0
        }

        for df in self.results.get('dataflows', []):
            score = df.get('TransformationScore', 0)
            try:
                score = float(score)
            except:
                score = 0

            if score >= 100:
                complexity_distribution['Critical (100+)'] += 1
            elif score >= 50:
                complexity_distribution['High (50-99)'] += 1
            elif score >= 20:
                complexity_distribution['Medium (20-49)'] += 1
            else:
                complexity_distribution['Low (<20)'] += 1

        total = sum(complexity_distribution.values())

        colors = {
            'Critical (100+)': 'C00000',
            'High (50-99)': 'FF6600',
            'Medium (20-49)': 'FFC000',
            'Low (<20)': '92D050'
        }

        for level, count in complexity_distribution.items():
            pct = (count / total * 100) if total > 0 else 0

            ws.cell(start_row, 1).value = level
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

            ws.cell(start_row, 2).value = count
            ws.cell(start_row, 2).font = Font(bold=True, size=11)
            ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

            bar_length = int(pct / 5)
            bar_cell = ws.cell(start_row, 3)
            bar_cell.value = "‚ñà" * bar_length
            bar_cell.font = Font(size=14, color=colors[level])

            ws.cell(start_row, 4).value = f"{pct:.1f}%"
            ws.cell(start_row, 4).font = Font(size=10)
            ws.cell(start_row, 4).fill = PatternFill(start_color=colors[level], end_color=colors[level], fill_type='solid')
            ws.cell(start_row, 4).font = Font(bold=True, color='FFFFFF')
            ws.cell(start_row, 4).alignment = Alignment(horizontal='center')

            start_row += 1

        return start_row + 1

    def _write_performance_insights(self, ws, start_row: int) -> int:
        """
         PERFORMANCE INSIGHTS & BOTTLENECK DETECTION
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "‚ö° PERFORMANCE INSIGHTS & BOTTLENECK DETECTION"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='9900CC', end_color='9900CC', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        bottlenecks = []

        large_pipelines = [
            p for p in self.results['pipeline_analysis']
            if p.get('TotalActivities', 0) > 50
        ]

        if large_pipelines:
            bottlenecks.append({
                'type': ' Large Pipelines',
                'count': len(large_pipelines),
                'description': f'{len(large_pipelines)} pipelines with >50 activities',
                'impact': 'Long execution times',
                'recommendation': 'Consider splitting into smaller pipelines'
            })

        deep_nesting = [
            p for p in self.results['pipeline_analysis']
            if p.get('MaxNestingDepth', 0) > 5
        ]

        if deep_nesting:
            bottlenecks.append({
                'type': 'üîÑ Deep Nesting',
                'count': len(deep_nesting),
                'description': f'{len(deep_nesting)} pipelines with nesting depth >5',
                'impact': 'Complex debugging, maintenance issues',
                'recommendation': 'Flatten control flow structures'
            })

        auto_resolve_count = len([
            a for a in self.results['activities']
            if a.get('IntegrationRuntime') == 'AutoResolveIR'
        ])

        if auto_resolve_count > 100:
            bottlenecks.append({
                'type': 'üåê AutoResolve IR',
                'count': auto_resolve_count,
                'description': f'{auto_resolve_count} activities using AutoResolveIR',
                'impact': 'Unpredictable performance',
                'recommendation': 'Specify dedicated Integration Runtimes'
            })

        sequential_pipelines = [
            p for p in self.results['pipeline_analysis']
            if p.get('LoopActivities', 0) > 0 and p.get('TotalActivities', 0) > 20
        ]

        if sequential_pipelines:
            bottlenecks.append({
                'type': 'üêå Sequential Processing',
                'count': len(sequential_pipelines),
                'description': f'{len(sequential_pipelines)} pipelines may benefit from parallelization',
                'impact': 'Slow overall execution',
                'recommendation': 'Use ForEach with parallel execution'
            })

        if bottlenecks:

            headers = ['Bottleneck Type', 'Count', 'Impact', 'Recommendation']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(start_row, col)
                cell.value = header
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.border = ExcelBorders.thin_border()
                cell.alignment = Alignment(horizontal='center')

            start_row += 1

            for bottleneck in bottlenecks:
                ws.cell(start_row, 1).value = bottleneck['type']
                ws.cell(start_row, 1).font = Font(bold=True, size=10)

                ws.cell(start_row, 2).value = bottleneck['count']
                ws.cell(start_row, 2).font = Font(bold=True, color='CC0000')
                ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

                ws.cell(start_row, 3).value = bottleneck['impact']
                ws.cell(start_row, 3).font = Font(size=9)

                ws.cell(start_row, 4).value = bottleneck['recommendation']
                ws.cell(start_row, 4).font = Font(size=9, italic=True)

                for col in range(1, 5):
                    ws.cell(start_row, col).border = ExcelBorders.thin_border()

                start_row += 1
        else:
            ws.cell(start_row, 1).value = " No significant performance bottlenecks detected!"
            ws.merge_cells(f'A{start_row}:D{start_row}')
            ws.cell(start_row, 1).font = Font(bold=True, color='00B050', size=11)
            start_row += 1

        return start_row + 1

    def _write_top_pipelines_ranking(self, ws, start_row: int) -> int:
        """
         TOP PIPELINES RANKING

        Shows:
        - Most complex pipelines
        - Highest impact pipelines
        - Most active pipelines
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " TOP PIPELINES RANKING"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        ws.cell(start_row, 1).value = "üî• Most Complex Pipelines"
        ws.cell(start_row, 1).font = Font(bold=True, size=11, color='C00000')
        ws.merge_cells(f'A{start_row}:D{start_row}')
        start_row += 1

        ws.cell(start_row, 1).value = "Rank"
        ws.cell(start_row, 2).value = "Pipeline"
        ws.cell(start_row, 3).value = "Complexity"
        ws.cell(start_row, 4).value = "Activities"

        for col in range(1, 5):
            ws.cell(start_row, col).font = Font(bold=True, size=9)
            ws.cell(start_row, col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        start_row += 1

        sorted_pipelines = sorted(
            self.results['pipeline_analysis'],
            key=lambda p: p.get('ComplexityScore', 0),
            reverse=True
        )[:10]

        for rank, pipeline in enumerate(sorted_pipelines, 1):
            medal = "ü•á" if rank == 1 else "ü•à" if rank == 2 else "ü•â" if rank == 3 else f"{rank}."

            ws.cell(start_row, 1).value = medal
            ws.cell(start_row, 1).alignment = Alignment(horizontal='center')

            ws.cell(start_row, 2).value = pipeline['Pipeline']
            ws.cell(start_row, 2).font = Font(size=9)

            ws.cell(start_row, 3).value = pipeline.get('ComplexityScore', 0)
            ws.cell(start_row, 3).font = Font(bold=True, color='C00000')
            ws.cell(start_row, 3).alignment = Alignment(horizontal='center')

            ws.cell(start_row, 4).value = pipeline.get('TotalActivities', 0)
            ws.cell(start_row, 4).alignment = Alignment(horizontal='center')

            start_row += 1

        start_row += 1

        ws.cell(start_row, 1).value = "üí• Highest Impact Pipelines"
        ws.cell(start_row, 1).font = Font(bold=True, size=11, color='FF6600')
        ws.merge_cells(f'A{start_row}:D{start_row}')
        start_row += 1

        ws.cell(start_row, 1).value = "Rank"
        ws.cell(start_row, 2).value = "Pipeline"
        ws.cell(start_row, 3).value = "Impact"
        ws.cell(start_row, 4).value = "Blast Radius"

        for col in range(1, 5):
            ws.cell(start_row, col).font = Font(bold=True, size=9)
            ws.cell(start_row, col).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        start_row += 1

        impact_order = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
        sorted_impact = sorted(
            self.results['impact_analysis'],
            key=lambda p: (impact_order.get(p.get('Impact', 'LOW'), 99), -p.get('BlastRadius', 0))
        )[:10]

        for rank, pipeline in enumerate(sorted_impact, 1):
            medal = "ü•á" if rank == 1 else "ü•à" if rank == 2 else "ü•â" if rank == 3 else f"{rank}."

            ws.cell(start_row, 1).value = medal
            ws.cell(start_row, 1).alignment = Alignment(horizontal='center')

            ws.cell(start_row, 2).value = pipeline['Pipeline']
            ws.cell(start_row, 2).font = Font(size=9)

            impact = pipeline.get('Impact', 'UNKNOWN')
            ws.cell(start_row, 3).value = impact
            ws.cell(start_row, 3).font = Font(bold=True)
            ws.cell(start_row, 3).alignment = Alignment(horizontal='center')

            if impact == 'CRITICAL':
                ws.cell(start_row, 3).fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                ws.cell(start_row, 3).font = Font(bold=True, color='FFFFFF')
            elif impact == 'HIGH':
                ws.cell(start_row, 3).fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
                ws.cell(start_row, 3).font = Font(bold=True, color='FFFFFF')

            ws.cell(start_row, 4).value = pipeline.get('BlastRadius', 0)
            ws.cell(start_row, 4).alignment = Alignment(horizontal='center')

            start_row += 1

        return start_row + 1

    def _write_security_compliance_checklist(self, ws, start_row: int) -> int:
        """
         SECURITY & COMPLIANCE CHECKLIST

        Best practices assessment:
        - Key Vault usage
        - Integration Runtime security
        - Managed Identity
        - Network security
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üîí SECURITY & COMPLIANCE CHECKLIST"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        checks = []

        kv_usage = len([
            ls for ls in self.results['linked_services']
            if ls.get('UsesKeyVault') == 'Yes'
        ])
        total_ls = len(self.results['linked_services'])
        kv_percentage = (kv_usage / total_ls * 100) if total_ls > 0 else 0

        checks.append({
            'check': 'Key Vault Integration',
            'status': ' PASS' if kv_percentage > 50 else ' REVIEW',
            'detail': f'{kv_usage}/{total_ls} ({kv_percentage:.0f}%) linked services use Key Vault',
            'recommendation': 'Good practice' if kv_percentage > 50 else 'Consider using Key Vault for secrets'
        })

        mi_usage = len([
            ls for ls in self.results['linked_services']
            if 'Managed Identity' in ls.get('Authentication', '')
        ])
        mi_percentage = (mi_usage / total_ls * 100) if total_ls > 0 else 0

        checks.append({
            'check': 'Managed Identity Usage',
            'status': ' PASS' if mi_percentage > 30 else ' REVIEW',
            'detail': f'{mi_usage}/{total_ls} ({mi_percentage:.0f}%) use Managed Identity',
            'recommendation': 'Good security practice' if mi_percentage > 30 else 'Consider Managed Identity for Azure resources'
        })

        self_hosted_ir = len([
            ir for ir in self.results['integration_runtimes']
            if ir.get('Type') == 'SelfHosted'
        ])

        checks.append({
            'check': 'Self-Hosted IR Security',
            'status': '‚Ñπ INFO',
            'detail': f'{self_hosted_ir} self-hosted IRs detected',
            'recommendation': 'Ensure network security and patching for self-hosted IRs'
        })

        vnet_irs = len([
            ir for ir in self.results['integration_runtimes']
            if ir.get('VNetIntegration') == 'Yes'
        ])

        checks.append({
            'check': 'VNet Integration',
            'status': ' PASS' if vnet_irs > 0 else '‚Ñπ INFO',
            'detail': f'{vnet_irs} IRs with VNet integration',
            'recommendation': 'VNet integration enhances security' if vnet_irs > 0 else 'Consider VNet integration for sensitive data'
        })

        has_credentials = len(self.results.get('credentials', [])) > 0

        checks.append({
            'check': 'Credential Management',
            'status': ' PASS' if has_credentials else '‚Ñπ INFO',
            'detail': f"{len(self.results.get('credentials', []))} managed credentials",
            'recommendation': 'Good practice' if has_credentials else 'Consider using ADF Credentials for centralized auth'
        })

        for check in checks:
            ws.cell(start_row, 1).value = check['check']
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

            ws.cell(start_row, 2).value = check['status']
            ws.cell(start_row, 2).font = Font(bold=True, size=10)
            ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

            if '' in check['status']:
                ws.cell(start_row, 2).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            elif '' in check['status']:
                ws.cell(start_row, 2).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

            ws.cell(start_row, 3).value = check['detail']
            ws.cell(start_row, 3).font = Font(size=9)

            ws.cell(start_row, 4).value = check['recommendation']
            ws.cell(start_row, 4).font = Font(size=9, italic=True)

            for col in range(1, 5):
                ws.cell(start_row, col).border = ExcelBorders.thin_border()

            start_row += 1

        return start_row + 1

    def _write_activity_distribution_chart(self, ws, start_row: int) -> int:
        """
         ACTIVITY TYPE DISTRIBUTION

        Visual chart of activity type usage
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " ACTIVITY TYPE DISTRIBUTION"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        top_activities = self.metrics['activity_types'].most_common(15)
        total_activities = sum(self.metrics['activity_types'].values())

        for activity_type, count in top_activities:
            percentage = (count / total_activities * 100) if total_activities > 0 else 0

            ws.cell(start_row, 1).value = activity_type
            ws.cell(start_row, 1).font = Font(size=9)

            ws.cell(start_row, 2).value = count
            ws.cell(start_row, 2).font = Font(bold=True, size=10)
            ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

            bar_length = int(percentage / 2)  # Scale
            ws.cell(start_row, 3).value = "‚ñà" * bar_length
            ws.cell(start_row, 3).font = Font(size=12, color='4472C4')

            ws.cell(start_row, 4).value = f"{percentage:.1f}%"
            ws.cell(start_row, 4).font = Font(size=9)

            start_row += 1

        return start_row + 1

    def _write_data_flow_network_stats(self, ws, start_row: int) -> int:
        """
         DATA FLOW NETWORK STATISTICS
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = "üåê DATA FLOW NETWORK STATISTICS"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        total_nodes = len(self.graph)
        total_edges = sum(len(node_data['depends_on']) for node_data in self.graph.values())

        most_connected = sorted(
            [(node, len(data['depends_on']) + len(data['used_by']))
             for node, data in self.graph.items()],
            key=lambda x: x[1],
            reverse=True
        )[:5]

        metrics = [
            ("Total Nodes (Resources)", total_nodes, "All resources in dependency graph"),
            ("Total Edges (Dependencies)", total_edges, "Direct dependency relationships"),
            ("Avg Connections per Node", f"{total_edges / total_nodes:.1f}" if total_nodes > 0 else "0", "Network density indicator"),
            ("Isolated Resources", len([n for n, d in self.graph.items() if not d['depends_on'] and not d['used_by']]), "Resources with no connections"),
        ]

        for metric, value, description in metrics:
            ws.cell(start_row, 1).value = metric
            ws.cell(start_row, 1).font = Font(bold=True, size=10)

            ws.cell(start_row, 2).value = value
            ws.cell(start_row, 2).font = Font(bold=True, size=11, color='0066CC')
            ws.cell(start_row, 2).alignment = Alignment(horizontal='center')

            ws.cell(start_row, 3).value = description
            ws.merge_cells(f'C{start_row}:D{start_row}')
            ws.cell(start_row, 3).font = Font(size=9, italic=True)

            start_row += 1

        start_row += 1
        ws.cell(start_row, 1).value = "Most Connected Resources:"
        ws.cell(start_row, 1).font = Font(bold=True, size=10)
        ws.merge_cells(f'A{start_row}:D{start_row}')
        start_row += 1

        for node, connections in most_connected:
            ws.cell(start_row, 1).value = f"‚Ä¢ {node}"
            ws.cell(start_row, 2).value = f"{connections} connections"
            ws.merge_cells(f'A{start_row}:C{start_row}')
            ws.cell(start_row, 1).font = Font(size=9)
            start_row += 1

        return start_row + 1

    def _write_change_risk_assessment(self, ws, start_row: int) -> int:
        """
         CHANGE RISK ASSESSMENT
        """

        header_cell = ws.cell(start_row, 1)
        header_cell.value = " CHANGE RISK ASSESSMENT"
        ws.merge_cells(f'A{start_row}:D{start_row}')

        header_cell.font = Font(size=14, bold=True, color='FFFFFF')
        header_cell.fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='left', vertical='center')

        start_row += 1

        risks = [
            {
                'category': ' High Risk Changes',
                'resources': [p['Pipeline'] for p in self.results['impact_analysis'] if p.get('Impact') == 'CRITICAL'][:5],
                'description': 'Changes to these pipelines affect many dependencies',
                'mitigation': 'Thorough testing, staged rollout, backup plan'
            },
            {
                'category': ' Medium Risk Changes',
                'resources': [p['Pipeline'] for p in self.results['impact_analysis'] if p.get('Impact') == 'HIGH'][:5],
                'description': 'Significant but contained impact',
                'mitigation': 'Standard testing, monitor closely'
            },
            {
                'category': ' Low Risk Changes',
                'resources': [p['Pipeline'] for p in self.results['impact_analysis'] if p.get('Impact') == 'LOW'][:5],
                'description': 'Isolated or orphaned resources',
                'mitigation': 'Basic testing sufficient'
            }
        ]

        for risk in risks:

            ws.cell(start_row, 1).value = risk['category']
            ws.cell(start_row, 1).font = Font(bold=True, size=11)
            ws.merge_cells(f'A{start_row}:D{start_row}')

            if '' in risk['category']:
                ws.cell(start_row, 1).fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            elif '' in risk['category']:
                ws.cell(start_row, 1).fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif '' in risk['category']:
                ws.cell(start_row, 1).fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')

            start_row += 1

            ws.cell(start_row, 1).value = f"Count: {len(risk['resources'])}"
            ws.cell(start_row, 1).font = Font(size=9)
            start_row += 1

            if risk['resources']:
                ws.cell(start_row, 1).value = "Examples:"
                ws.cell(start_row, 1).font = Font(size=9, italic=True)
                start_row += 1

                for resource in risk['resources'][:3]:
                    ws.cell(start_row, 1).value = f"  ‚Ä¢ {resource}"
                    ws.cell(start_row, 1).font = Font(size=8)
                    start_row += 1

            ws.cell(start_row, 1).value = f"Mitigation: {risk['mitigation']}"
            ws.cell(start_row, 1).font = Font(size=9, italic=True, color='666666')
            ws.merge_cells(f'A{start_row}:D{start_row}')
            start_row += 2

        return start_row

    def _calculate_quality_score(self) -> int:
        """Calculate quality score (0-100)"""
        score = 100

        circular_deps = len(self.results['circular_dependencies'])
        score -= min(circular_deps * 10, 30)

        orphaned = (
            len(self.results['orphaned_pipelines']) +
            len(self.results['orphaned_datasets']) +
            len(self.results['orphaned_linked_services'])
        )
        orphan_percentage = (orphaned / max(len(self.resources['all']), 1)) * 100
        score -= min(orphan_percentage, 20)

        broken_triggers = len([t for t in self.results['orphaned_triggers'] if t.get('Type') == 'BrokenReference'])
        score -= min(broken_triggers * 5, 15)

        return max(0, min(100, int(score)))

    def _calculate_performance_score(self) -> int:
        """Calculate performance score (0-100)"""
        score = 100

        complex_pipelines = len([
            p for p in self.results['pipeline_analysis']
            if p.get('ComplexityScore', 0) > 100
        ])
        total_pipelines = len(self.results['pipeline_analysis'])
        if total_pipelines > 0:
            complex_percentage = (complex_pipelines / total_pipelines) * 100
            score -= min(complex_percentage, 25)

        deep_nesting = len([
            p for p in self.results['pipeline_analysis']
            if p.get('MaxNestingDepth', 0) > 5
        ])
        if total_pipelines > 0:
            nesting_percentage = (deep_nesting / total_pipelines) * 100
            score -= min(nesting_percentage, 15)

        auto_resolve = len([
            a for a in self.results['activities']
            if a.get('IntegrationRuntime') == 'AutoResolveIR'
        ])
        total_activities = len(self.results['activities'])
        if total_activities > 0:
            auto_percentage = (auto_resolve / total_activities) * 100
            score -= min(auto_percentage / 2, 10)

        return max(0, min(100, int(score)))

    def _calculate_security_score(self) -> int:
        """Calculate security score (0-100)"""
        score = 100

        kv_usage = len([
            ls for ls in self.results['linked_services']
            if ls.get('UsesKeyVault') == 'Yes'
        ])
        total_ls = len(self.results['linked_services'])
        if total_ls > 0:
            kv_percentage = (kv_usage / total_ls) * 100
            if kv_percentage < 50:
                score -= (50 - kv_percentage) / 2

        mi_usage = len([
            ls for ls in self.results['linked_services']
            if 'Managed Identity' in ls.get('Authentication', '')
        ])
        if total_ls > 0:
            mi_percentage = (mi_usage / total_ls) * 100
            if mi_percentage < 30:
                score -= (30 - mi_percentage) / 2

        vnet_irs = len([
            ir for ir in self.results['integration_runtimes']
            if ir.get('VNetIntegration') == 'Yes'
        ])
        if vnet_irs == 0 and len(self.results['integration_runtimes']) > 0:
            score -= 10

        return max(0, min(100, int(score)))

    def _calculate_maintainability_score(self) -> int:
        """Calculate maintainability score (0-100)"""
        score = 100

        poorly_named = len([
            p for p in self.results['pipelines']
            if len(p.get('Pipeline', '')) < 5 or not any(c.isupper() for c in p.get('Pipeline', ''))
        ])
        total_pipelines = len(self.results['pipelines'])
        if total_pipelines > 0:
            poorly_named_percentage = (poorly_named / total_pipelines) * 100
            score -= min(poorly_named_percentage / 2, 15)

        no_description = len([
            p for p in self.results['pipelines']
            if not p.get('Description')
        ])
        if total_pipelines > 0:
            no_desc_percentage = (no_description / total_pipelines) * 100
            score -= min(no_desc_percentage / 3, 10)

        no_folder = len([
            p for p in self.results['pipelines']
            if not p.get('Folder')
        ])
        if total_pipelines > 0:
            no_folder_percentage = (no_folder / total_pipelines) * 100
            score -= min(no_folder_percentage / 3, 10)

        return max(0, min(100, int(score)))

    def _get_health_color(self, score: int) -> str:
        """Get color for health score"""
        if score >= 80:
            return '00B050'  # Green
        elif score >= 60:
            return 'FFC000'  # Yellow
        elif score >= 40:
            return 'FF6600'  # Orange
        else:
            return 'C00000'  # Red

    def _get_health_status(self, score: int) -> str:
        """Get health status text"""
        if score >= 90:
            return "üåü EXCELLENT"
        elif score >= 80:
            return " GOOD"
        elif score >= 60:
            return " FAIR"
        elif score >= 40:
            return "üî∂ NEEDS IMPROVEMENT"
        else:
            return " CRITICAL"

    analyzer_class._write_health_score_dashboard = _write_health_score_dashboard

    if EnhancementConfig.is_enabled(ENHANCEMENT_CONFIG, 'advanced_dashboard', 'cost_analysis'):
        analyzer_class._write_cost_analysis = _write_cost_analysis
    analyzer_class._write_complexity_heat_map = _write_complexity_heat_map
    analyzer_class._write_performance_insights = _write_performance_insights
    analyzer_class._write_top_pipelines_ranking = _write_top_pipelines_ranking
    analyzer_class._write_security_compliance_checklist = _write_security_compliance_checklist
    analyzer_class._write_activity_distribution_chart = _write_activity_distribution_chart
    analyzer_class._write_data_flow_network_stats = _write_data_flow_network_stats
    analyzer_class._write_dataflow_complexity_heat_map = _write_dataflow_complexity_heat_map
    analyzer_class._write_change_risk_assessment = _write_change_risk_assessment

    analyzer_class._calculate_quality_score = _calculate_quality_score
    analyzer_class._calculate_performance_score = _calculate_performance_score
    analyzer_class._calculate_security_score = _calculate_security_score
    analyzer_class._calculate_maintainability_score = _calculate_maintainability_score
    analyzer_class._get_health_color = _get_health_color
    analyzer_class._get_health_status = _get_health_status

    print("   Advanced summary sections applied")

def integrate_advanced_sections_into_summary(analyzer_class):
    """
     Integrate advanced sections into enhanced summary sheet
    """

    def _write_complete_enhanced_summary_sheet(self, writer, timestamp: str):
        """
         COMPLETE ENHANCED SUMMARY WITH ADVANCED SECTIONS
        """

        import pandas as pd

        workbook = writer.book

        df_init = pd.DataFrame({'_': ['']})
        sheet_name = self._get_unique_sheet_name('Summary')
        df_init.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        ws = writer.sheets[sheet_name]

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50

        current_row = 1

        current_row = self._write_project_banner(ws, current_row, timestamp)
        current_row += 2
        current_row = self._write_executive_summary(ws, current_row)
        current_row += 2
        current_row = self._write_critical_alerts(ws, current_row)

        current_row += 2
        current_row = self._write_health_score_dashboard(ws, current_row)

        if EnhancementConfig.is_enabled(ENHANCEMENT_CONFIG, 'advanced_dashboard', 'cost_analysis') and hasattr(self, '_write_cost_analysis'):
            current_row += 2
            current_row = self._write_cost_analysis(ws, current_row)

        current_row += 2
        current_row = self._write_complexity_heat_map(ws, current_row)

        # Add DataFlow complexity heat map (uses TransformationScore)
        current_row += 1
        current_row = self._write_dataflow_complexity_heat_map(ws, current_row)

        current_row += 2
        current_row = self._write_performance_insights(ws, current_row)

        current_row += 2
        current_row = self._write_top_pipelines_ranking(ws, current_row)

        current_row += 2
        current_row = self._write_security_compliance_checklist(ws, current_row)

        current_row += 2
        current_row = self._write_activity_distribution_chart(ws, current_row)

        current_row += 2
        current_row = self._write_data_flow_network_stats(ws, current_row)

        current_row += 2
        current_row = self._write_change_risk_assessment(ws, current_row)

        current_row += 2
        current_row = self._write_metrics_dashboard(ws, current_row)
        current_row += 2
        current_row = self._write_resource_overview(ws, current_row)
        current_row += 2
        current_row = self._write_recommendations(ws, current_row)
        current_row += 2
        current_row = self._write_detailed_statistics(ws, current_row, timestamp)

        self.logger.info(f"  ‚úì Complete Enhanced Summary with Advanced Sections")

    analyzer_class._write_summary_sheet = _write_complete_enhanced_summary_sheet

    print("   Advanced sections integrated into summary sheet")

def apply_complete_excel_enhancements(analyzer_class=None, verbose: bool = True):
    """
     ULTIMATE FUNCTION: Apply ALL enhancements including advanced sections

    Usage:
        from adf_analyzer_v10_excel_enhancements import apply_complete_excel_enhancements

        apply_complete_excel_enhancements()
    """

    if analyzer_class is None:
        try:
            from adf_analyzer_v10_complete import UltimateEnterpriseADFAnalyzer
            analyzer_class = UltimateEnterpriseADFAnalyzer
        except ImportError:
            print(" ERROR: Could not import analyzer")
            return False

    if verbose:
        print("\n" + "="*80)
        print(" APPLYING COMPLETE EXCEL ENHANCEMENTS (ULTIMATE EDITION)")
        print("="*80 + "\n")

    try:

        if verbose:
            print("üì¶ Parts 1-4: Base formatting, conditional formatting, hyperlinks...")

        create_enhanced_export_function(analyzer_class)
        create_enhanced_beautification_method(analyzer_class)

        if verbose:
            print("üì¶ Part 5: Enhanced summary sheet...")

        create_enhanced_summary_sheet_writer(analyzer_class)

        if verbose:
            print("üì¶ Part 6: Advanced dashboard sections...")

        add_advanced_summary_sections(analyzer_class)
        integrate_advanced_sections_into_summary(analyzer_class)

        if verbose:
            print("\n" + "="*80)
            print(" COMPLETE EXCEL ENHANCEMENTS APPLIED (ULTIMATE EDITION)")
            print("="*80)
            print("\nüé® Summary Sheet Now Includes:")
            print("   Beautiful project banner")
            print("   Executive summary")
            print("   Critical alerts")
            print("   üè• Health Score Dashboard (Quality, Performance, Security)")
            if EnhancementConfig.is_enabled(ENHANCEMENT_CONFIG, 'advanced_dashboard', 'cost_analysis'):
                print("   üí∞ Cost Analysis & Optimization")
            print("   üå° Complexity Heat Map")
            print("   ‚ö° Performance Insights & Bottlenecks")
            print("    Top Pipelines Ranking")
            print("   üîí Security & Compliance Checklist")
            print("    Activity Distribution Chart")
            print("   üåê Data Flow Network Statistics")
            print("    Change Risk Assessment")
            print("    Recommendations")
            print("    Detailed Metrics")
            print("   üîó Navigation Links")
            print("\nüåü Plus ALL formatting enhancements from Parts 1-4!")
            print("="*80 + "\n")

        return True

    except Exception as e:
        if verbose:
            print(f"\n FAILED: {e}")
            traceback.print_exc()
        return False

print(" Part 6/6 loaded: Advanced Summary Enhancements loaded")
print("\n" + "="*80)
print("üéâ ALL 6 PARTS LOADED SUCCESSFULLY!")
print("="*80)
print("\nüìö For usage guide, run:")
print("   from adf_analyzer_v10_excel_enhancements import print_usage_guide")
print("   print_usage_guide()")
print("\nüöÄ To apply enhancements:")
print("   from adf_analyzer_v10_excel_enhancements import apply_excel_enhancements")
print("   Use: apply_complete_excel_enhancements() for ULTIMATE enhancement")
print("="*80 + "\n")